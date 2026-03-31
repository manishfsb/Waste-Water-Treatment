"""
Extract lab report data from Excel files (2022-2025) into JSON,
then generate a consolidated verification Excel file per year.

Structural notes:
- 2022/2023: two inlet sections (composite cols 7-12, non-composite cols 13-16)
             two CCT sections (composite cols 38-43, non-composite cols 44-47)
             Power in MW
- 2024/2025: non-composite inlet cols 7-10, composite inlet cols 12-20
             non-composite CCT cols 49-52, composite CCT cols 40-47
             Power in KWH
We consistently use the NON-COMPOSITE (grab) section for all years,
which is analogous to the single section that existed in 2021.
"""

import json
import os
import re
from datetime import datetime

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_START_ROW = 9
CONTROL_LIMIT_ROW = 7

# ── Per-year column configurations ───────────────────────────────────────────

YEAR_CONFIGS = {
    2022: {
        "months": {
            "January":   "Lab Report January-2022.xlsx",
            "February":  "Lab Report February-2022.xlsx",
            "March":     "Lab Report March-2022.xlsx",
            "April":     "Lab Report April-2022.xlsx",
            "May":       "LAB REPORT may 2022.xlsx",
            "June":      "Lab Report June-2022.xlsx",
            "July":      "Lab Report -July  2022.xlsx",
            "August":    "Lab Report -August 2022.xlsx",
            "September": "Lab Report -Septmeber  2022.xlsx",
            "October":   "Lab Report -October 2022.xlsx",
            "November":  "Lab Report -November 2022.xlsx",
            "December":  "Lab Report -Decemeber  2022 .xlsx",
        },
    },
    2023: {
        "months": {
            "January":   "Lab Report -January 2023.xlsx",
            "February":  "Lab Report -February  2023.xlsx",
            "March":     "Lab Report -March 2023.xlsx",
            "April":     "Lab Report -April month 2023.xlsx",
            "May":       "LAB REPORT MAY 2023.xlsx",
            "June":      "Lab Report June  2023.xlsx",
            "July":      "LAB REPORT JULY 2023.xlsx",
            "August":    "LAB REPORT AUGUST 2023.xlsx",
            "September": "LAB REPORT September 2023.xlsx",
            "October":   "LAB REPORT October 2023.xlsx",
            "November":  "LAB REPORT November 2023.xlsx",
            "December":  "LAB REPORT December 2023.xlsx",
        },
    },
    2024: {
        "months": {
            "January":   "LAB REPORT January 2024.xlsx",
            "February":  "LAB REPORT February 2024.xlsx",
            "March":     "LAB REPORT March 2024.xlsx",
            "April":     "Lab report for April 2024.xlsx",
            "May":       "LAB REPORT May 2024.xlsx",
            "June":      "LAB REPORT June 2024.xlsx",
            "July":      "LAB REPORT July 2024.xlsx",
            "August":    "LAB REPORT August 2024.xlsx",
            "September": "LAB REPORT September 2024.xlsx",
            "October":   "LAB REPORT October 2024.xlsx",
            "November":  "LAB REPORT November 2024.xlsx",
            "December":  "LAB REPORT December 2024.xlsx",
        },
    },
    2025: {
        "months": {
            "January":   "LAB REPORT January 2025.xlsx",
            "February":  "LAB REPORT February 2025.xlsx",
            "March":     "LAB REPORT March 2025.xlsx",
            "April":     "LAB REPORT April 2025.xlsx",
            "May":       "LAB REPORT MAY 2025.xlsx",
            "June":      "LAB REPORT JUNE 2025.xlsx",
            "July":      "LAB REPORT JULY 2025.xlsx",
            "August":    "LAB REPORT AUGUST 2025.xlsx",
            "September": "LAB REPORT SEPT 2025.xlsx",
            "October":   "LAB REPORT OCT 2025.xlsx",
            "November":  "LAB REPORT NOV 2025.xlsx",
            "December":  "LAB REPORT DEC 2025.xlsx",
        },
    },
}


def detect_columns(ws):
    """
    Auto-detect column positions from sheet headers.
    Returns (data_cols dict, ctrl_cols dict, power_unit string).

    Strategy:
    - Inlet grab section: find "Raw Sewage" (no "Composite", no "Flow") in row 5.
      pH/BOD/COD/TSS follow in row 6 from that column.
    - CCT non-composite: find "CHLORINE CONTACT TANK" (no "COMPOSITE","OUTLET","FLOW") in row 4.
      pH/BOD/COD/TSS/FRC follow in row 6 from that column.
    - Power unit: "KWH" in row 8 col 2.
    """
    r4 = list(ws.iter_rows(min_row=4, max_row=4, values_only=True))[0]
    r5 = list(ws.iter_rows(min_row=5, max_row=5, values_only=True))[0]
    r6 = list(ws.iter_rows(min_row=6, max_row=6, values_only=True))[0]
    r7 = list(ws.iter_rows(min_row=7, max_row=7, values_only=True))[0]
    r8 = list(ws.iter_rows(min_row=8, max_row=8, values_only=True))[0]

    def col_val(row, col):  # 1-based
        return row[col - 1] if col <= len(row) else None

    # ── Power unit ────────────────────────────────────────────────────────
    power_unit = "KWH" if (r8[1] and "KWH" in str(r8[1]).upper()) else "MW"

    # ── Grab inlet section ───────────────────────────────────────────────
    # Find "Raw Sewage" (grab, not composite) section header in row 5
    grab_section_col = None
    for i, v in enumerate(r5):
        if v is None:
            continue
        s = str(v).upper()
        if "RAW SEWAGE" in s and "COMPOSITE" not in s and "FLOW" not in s:
            grab_section_col = i + 1
            break

    # From grab_section_col, find pH/BOD/COD/TSS in row 6
    inlet_ph = inlet_bod = inlet_cod = inlet_tss = None
    if grab_section_col:
        params = ["PH", "BOD", "COD", "TSS"]
        found = []
        for i in range(grab_section_col - 1, min(grab_section_col + 15, len(r6))):
            v = r6[i]
            if v is None:
                continue
            s = str(v).upper()
            if "PH" in s and inlet_ph is None:
                inlet_ph = i + 1
                found.append("ph")
            elif "BOD" in s and inlet_bod is None:
                inlet_bod = i + 1
                found.append("bod")
            elif "COD" in s and inlet_cod is None:
                inlet_cod = i + 1
                found.append("cod")
            elif "TSS" in s and inlet_tss is None:
                inlet_tss = i + 1
                found.append("tss")
            if len(found) == 4:
                break

    # ── Non-composite CCT section ─────────────────────────────────────────
    cct_section_col = None
    for i, v in enumerate(r4):
        if v is None:
            continue
        s = str(v).upper()
        if "CHLORINE" in s and "COMPOSITE" not in s and "OUTLET" not in s and "FLOW" not in s:
            cct_section_col = i + 1
            break

    eff_ph = eff_bod = eff_cod = eff_tss = eff_frc = None
    if cct_section_col:
        found = []
        for i in range(cct_section_col - 1, min(cct_section_col + 15, len(r6))):
            v = r6[i]
            if v is None:
                continue
            s = str(v).upper()
            if "PH" in s and eff_ph is None:
                eff_ph = i + 1
                found.append("ph")
            elif "BOD" in s and eff_bod is None:
                eff_bod = i + 1
                found.append("bod")
            elif "COD" in s and eff_cod is None:
                eff_cod = i + 1
                found.append("cod")
            elif "TSS" in s and eff_tss is None:
                eff_tss = i + 1
                found.append("tss")
            elif "FRC" in s and eff_frc is None:
                eff_frc = i + 1
                found.append("frc")
            if len(found) == 5:
                break

    data_cols = {
        "power_ge":       2,
        "power_nea":      3,
        "power_total":    4,
        "power_per_flow": 5,
        "flow":           6,
        "inlet_ph":       inlet_ph,
        "inlet_bod":      inlet_bod,
        "inlet_cod":      inlet_cod,
        "inlet_tss":      inlet_tss,
        "effluent_ph":    eff_ph,
        "effluent_bod":   eff_bod,
        "effluent_cod":   eff_cod,
        "effluent_tss":   eff_tss,
        "effluent_frc":   eff_frc,
    }

    # Control limits sit in row 7 at the same columns as the data
    ctrl_cols = {
        "flow":          6,
        "inlet_ph":      inlet_ph,
        "inlet_bod":     inlet_bod,
        "inlet_cod":     inlet_cod,
        "inlet_tss":     inlet_tss,
        "effluent_ph":   eff_ph,
        "effluent_bod":  eff_bod,
        "effluent_cod":  eff_cod,
        "effluent_tss":  eff_tss,
    }

    return data_cols, ctrl_cols, power_unit

# ── Value parsing ─────────────────────────────────────────────────────────────

def parse_value(raw):
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw.strftime("%Y-%m-%d")
    if isinstance(raw, (int, float)):
        return raw
    s = str(raw).strip()
    if s in ("_", "-", "", "BDL", "*BDL", "N/A", "n/a",
             "#DIV/0!", "#REF!", "#N/A", "#VALUE!", "#NAME?", "Nil"):
        return None
    # Scientific notation with × (e.g., "3.8×106")
    m = re.match(r"^([0-9.]+)\s*[×xX]\s*10\^?(\d+)$", s)
    if m:
        return float(m.group(1)) * (10 ** int(m.group(2)))
    m = re.match(r"^([0-9.]+)\s*[xX]\s*(\d+)$", s)
    if m:
        return float(m.group(1)) * (10 ** int(m.group(2)))
    try:
        return float(s)
    except ValueError:
        return None  # unrecognised string → null


def parse_ctrl(raw):
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return raw
    s = str(raw).strip()
    return None if s in ("_", "-", "", "N/A") else s


# ── JSON extraction ───────────────────────────────────────────────────────────

def extract_month(year, month_name, filename):
    year_dir = os.path.join(BASE_DIR, str(year))
    filepath = os.path.join(year_dir, filename)
    if not os.path.exists(filepath):
        print(f"    WARNING: {filename} not found – skipping.")
        return None

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]

    data_cols, ctrl_cols, power_unit = detect_columns(ws)

    # Validate detection
    missing = [k for k, v in data_cols.items() if v is None and k not in ("effluent_frc",)]
    if missing:
        print(f"    WARNING {month_name}: could not detect columns for {missing}")

    # Control limits
    control_limits = {"power_per_flow": "< 482.02"}
    for key, col in ctrl_cols.items():
        if col is None:
            continue
        raw = ws.cell(row=CONTROL_LIMIT_ROW, column=col).value
        v = parse_ctrl(raw)
        if v is not None:
            control_limits[key] = v

    # Daily data — scan until terminator (supports embedded "Average" rows)
    days = []
    row_idx = DATA_START_ROW
    while row_idx < DATA_START_ROW + 50:  # generous upper bound
        date_cell = ws.cell(row=row_idx, column=1).value
        row_idx += 1
        if date_cell is None:
            break
        if isinstance(date_cell, str):
            s = date_cell.strip().upper()
            if s in ("AVG", "REMARKS:", ""):
                break
            if s == "AVERAGE":
                continue   # weekly sub-total row in some files – skip silently
            break           # any other string → end of data
        if not isinstance(date_cell, datetime):
            break

        day = {"date": date_cell.strftime("%Y-%m-%d")}
        for key, col in data_cols.items():
            if col is None:
                day[key] = None
            else:
                day[key] = parse_value(ws.cell(row=row_idx - 1, column=col).value)
        day["power_unit"] = power_unit   # store per-day so we know the unit
        days.append(day)

    wb.close()
    return {
        "month": month_name,
        "year": year,
        "power_unit": power_unit,
        "control_limits": control_limits,
        "days": days,
        # Store detected columns for verification
        "_data_cols": data_cols,
    }


def run_extraction(year):
    cfg = YEAR_CONFIGS[year]
    year_dir = os.path.join(BASE_DIR, str(year))
    data_dir = os.path.join(year_dir, "data")
    os.makedirs(data_dir, exist_ok=True)

    print(f"\n  Extracting {year}...")
    all_months = {}
    for month_name, filename in cfg["months"].items():
        data = extract_month(year, month_name, filename)
        if data is None:
            continue
        mk = month_name.lower()
        out_path = os.path.join(data_dir, f"{mk}.json")
        with open(out_path, "w") as f:
            json.dump(data, f, indent=2)
        print(f"    {month_name}: {len(data['days'])} days → {mk}.json")
        all_months[mk] = data

    combined = os.path.join(data_dir, "all_months.json")
    with open(combined, "w") as f:
        json.dump(all_months, f, indent=2)
    print(f"    Combined → all_months.json ({len(all_months)} months)")
    return all_months


# ── Excel generation ──────────────────────────────────────────────────────────

DISPLAY_COLS = [
    ("Date",                                         "date"),
    ("Power Gen. from Gas Engine",                   "power_ge"),
    ("Daily Power Consumption from NEA",             "power_nea"),
    ("Total Power Consumption NEA+GE",               "power_total"),
    ("Total Power Flow (KWh/ML)",                    "power_per_flow"),
    ("Inlet: Raw Sewage Flow (MLD)",                 "flow"),
    ("Inlet: pH",                                    "inlet_ph"),
    ("Inlet: BOD (mg/L)",                            "inlet_bod"),
    ("Inlet: COD (mg/L)",                            "inlet_cod"),
    ("Inlet: TSS (mg/L)",                            "inlet_tss"),
    ("Effluent: pH",                                 "effluent_ph"),
    ("Effluent: BOD (mg/L)",                         "effluent_bod"),
    ("Effluent: COD (mg/L)",                         "effluent_cod"),
    ("Effluent: TSS (mg/L)",                         "effluent_tss"),
]

CTRL_KEYS = {
    "flow":          "flow",
    "inlet_ph":      "inlet_ph",
    "inlet_bod":     "inlet_bod",
    "inlet_cod":     "inlet_cod",
    "inlet_tss":     "inlet_tss",
    "effluent_ph":   "effluent_ph",
    "effluent_bod":  "effluent_bod",
    "effluent_cod":  "effluent_cod",
    "effluent_tss":  "effluent_tss",
    "power_per_flow":"power_per_flow",
}

def _side(s="thin"):
    return Side(style=s)

def thin_border():
    s = _side()
    return Border(left=s, right=s, top=s, bottom=s)

HDR_MAIN     = PatternFill("solid", fgColor="1F4E79")
HDR_POWER    = PatternFill("solid", fgColor="2E75B6")
HDR_INLET    = PatternFill("solid", fgColor="375623")
HDR_EFFL     = PatternFill("solid", fgColor="833C00")
CTRL_FILL    = PatternFill("solid", fgColor="FFF2CC")
MONTH_FILL   = PatternFill("solid", fgColor="D6E4F0")

WHITE_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
DARK_FONT   = Font(name="Calibri", bold=True, color="000000", size=10)
NORMAL_FONT = Font(name="Calibri", size=10)
CTRL_FONT   = Font(name="Calibri", bold=True, color="7F0000", size=9, italic=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left", vertical="center")

def col_fill(jkey):
    if jkey in ("power_ge", "power_nea", "power_total", "power_per_flow"):
        return HDR_POWER
    if jkey.startswith("inlet") or jkey == "flow":
        return HDR_INLET
    if jkey.startswith("effluent"):
        return HDR_EFFL
    return HDR_MAIN

def style_header(cell, jkey):
    cell.fill = col_fill(jkey)
    cell.font = WHITE_FONT
    cell.alignment = CENTER
    cell.border = thin_border()

def style_data(cell, is_date=False):
    cell.font = NORMAL_FONT
    cell.alignment = LEFT if is_date else CENTER
    cell.border = thin_border()

def style_ctrl(cell):
    cell.fill = CTRL_FILL
    cell.font = CTRL_FONT
    cell.alignment = CENTER
    cell.border = thin_border()

def write_header_row(ws, row, power_unit, col_offset=1):
    for ci, (label, jkey) in enumerate(DISPLAY_COLS):
        # Adjust power column units dynamically
        if jkey in ("power_ge", "power_nea", "power_total"):
            display = f"{label} ({power_unit})"
        else:
            display = label
        cell = ws.cell(row=row, column=col_offset + ci, value=display)
        style_header(cell, jkey)
    return row + 1

def write_ctrl_row(ws, row, control_limits, col_offset=1):
    lbl = ws.cell(row=row, column=col_offset, value="Control Limit")
    lbl.fill = CTRL_FILL; lbl.font = CTRL_FONT
    lbl.alignment = CENTER; lbl.border = thin_border()
    for ci, (_, jkey) in enumerate(DISPLAY_COLS[1:], start=1):
        ck = CTRL_KEYS.get(jkey)
        v = control_limits.get(ck, "—") if ck else "—"
        cell = ws.cell(row=row, column=col_offset + ci, value=v if v is not None else "—")
        style_ctrl(cell)
    return row + 1

def write_data_rows(ws, days, row, col_offset=1):
    for day in days:
        for ci, (_, jkey) in enumerate(DISPLAY_COLS):
            raw = day.get(jkey)
            if jkey == "date" and raw:
                parts = raw.split("-")
                from datetime import date as dtdate
                value = dtdate(int(parts[0]), int(parts[1]), int(parts[2]))
            elif raw is None:
                value = ""
            else:
                value = round(raw, 2) if isinstance(raw, float) else raw
            cell = ws.cell(row=row, column=col_offset + ci, value=value)
            style_data(cell, is_date=(jkey == "date"))
            if jkey == "date" and not isinstance(value, str):
                cell.number_format = "DD-MMM-YYYY"
        row += 1
    return row

def set_col_widths(ws, col_offset=1):
    widths = [14, 14, 14, 14, 14, 12, 8, 10, 10, 10, 8, 10, 10, 10]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(col_offset + i)].width = w

def write_month_banner(ws, label, row, num_cols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    cell = ws.cell(row=row, column=1, value=label)
    cell.fill = MONTH_FILL
    cell.font = Font(name="Calibri", bold=True, size=11, color="1F4E79")
    cell.alignment = CENTER
    cell.border = thin_border()
    return row + 1

def generate_excel(year, all_months_data, power_unit=None):
    year_dir = os.path.join(BASE_DIR, str(year))
    output_path = os.path.join(year_dir, f"{year}_Lab_Report_Extracted_Data.xlsx")

    months_order = [
        "january", "february", "march", "april", "may", "june",
        "july", "august", "september", "october", "november", "december"
    ]
    months_present = [m for m in months_order if m in all_months_data]
    month_labels = {k: k.capitalize() for k in months_present}

    num_cols = len(DISPLAY_COLS)
    first_month = months_present[0].capitalize()
    last_month  = months_present[-1].capitalize()

    wb = Workbook()
    wb.remove(wb.active)

    # ── All Months sheet ──────────────────────────────────────────────────────
    ws_all = wb.create_sheet(f"All Months ({year})")
    ws_all.freeze_panes = "B3"

    ws_all.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    tc = ws_all.cell(row=1, column=1,
        value=f"{year} Lab Report – Extracted Data ({first_month}–{last_month})")
    tc.font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    tc.fill = HDR_MAIN
    tc.alignment = CENTER
    ws_all.row_dimensions[1].height = 22

    current_row = 2
    for mk in months_present:
        data = all_months_data[mk]
        month_power_unit = data.get("power_unit", power_unit or "MW")
        current_row = write_month_banner(ws_all, month_labels[mk], current_row, num_cols)
        current_row = write_header_row(ws_all, current_row, month_power_unit)
        current_row = write_ctrl_row(ws_all, current_row, data["control_limits"])
        current_row = write_data_rows(ws_all, data["days"], current_row)
        current_row += 1  # blank separator

    set_col_widths(ws_all)

    # ── Per-month sheets ──────────────────────────────────────────────────────
    for mk in months_present:
        data = all_months_data[mk]
        month_power_unit = data.get("power_unit", power_unit or "MW")
        ws = wb.create_sheet(month_labels[mk])
        ws.freeze_panes = "B4"

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
        tc = ws.cell(row=1, column=1,
            value=f"{year} Lab Report – {month_labels[mk]} (Extracted Data)")
        tc.font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
        tc.fill = HDR_MAIN
        tc.alignment = CENTER
        ws.row_dimensions[1].height = 22

        row = 2
        row = write_header_row(ws, row, month_power_unit)
        row = write_ctrl_row(ws, row, data["control_limits"])
        write_data_rows(ws, data["days"], row)
        set_col_widths(ws)

    wb.save(output_path)
    print(f"    Excel saved: {os.path.basename(output_path)}")
    return output_path


# ── Verification ──────────────────────────────────────────────────────────────

def verify_chain(year, all_months_data):
    """Verify JSON values against source Excel files using auto-detected columns."""
    cfg = YEAR_CONFIGS[year]
    year_dir = os.path.join(BASE_DIR, str(year))

    total_days = 0
    total_errors = 0

    for month_name, filename in cfg["months"].items():
        mk = month_name.lower()
        if mk not in all_months_data:
            continue

        filepath = os.path.join(year_dir, filename)
        if not os.path.exists(filepath):
            continue

        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        ws = wb.active

        # Use the same auto-detected columns that were used during extraction
        data_cols = all_months_data[mk].get("_data_cols", {})
        if not data_cols:
            data_cols, _, _ = detect_columns(ws)

        json_days = {d["date"]: d for d in all_months_data[mk]["days"]}

        errors = []
        matched = 0

        for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
            date_val = row[0]
            if not isinstance(date_val, datetime):
                if isinstance(date_val, str) and date_val.strip().upper() == "AVERAGE":
                    continue
                continue
            date_str = date_val.strftime("%Y-%m-%d")
            if date_str not in json_days:
                errors.append(f"  {date_str}: in Excel but missing in JSON")
                continue
            matched += 1
            day = json_days[date_str]

            for key, col in data_cols.items():
                if col is None:
                    continue
                xl_raw = row[col - 1]
                xl_val = parse_value(xl_raw)
                json_val = day.get(key)

                def norm(v):
                    if v is None:
                        return None
                    return round(v, 2) if isinstance(v, float) else v

                xn, jn = norm(xl_val), norm(json_val)
                if xn != jn:
                    if isinstance(xn, float) and isinstance(jn, float):
                        if abs(xn - jn) < 0.01:
                            continue
                    errors.append(f"  {date_str} [{key}]: XL={xn!r}  JSON={jn!r}")

        wb.close()
        total_days += matched
        total_errors += len(errors)

        if errors:
            print(f"    {month_name}: {matched} days, {len(errors)} discrepancies:")
            for e in errors[:5]:
                print(f"      {e}")
            if len(errors) > 5:
                print(f"      ... and {len(errors)-5} more")
        else:
            print(f"    {month_name}: {matched} days — OK")

    print(f"  → {year} total: {total_days} days, {total_errors} discrepancies")
    return total_errors


# ── Main ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    grand_total_errors = 0

    for year in [2022, 2023, 2024, 2025]:
        print(f"\n{'='*60}")
        print(f"  YEAR {year}")
        print(f"{'='*60}")

        # Step 1: Extract to JSON
        all_months = run_extraction(year)

        # Step 2: Generate Excel (power_unit is per-month from JSON)
        generate_excel(year, all_months)

        # Step 3: Verify chain (source Excel → JSON)
        print(f"\n  Verifying chain (source Excel → JSON)...")
        errs = verify_chain(year, all_months)
        grand_total_errors += errs

    print(f"\n{'='*60}")
    print(f"ALL DONE – grand total discrepancies: {grand_total_errors}")
