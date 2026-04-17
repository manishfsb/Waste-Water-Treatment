"""
Extract lab report data from Excel files (2022-2025) into JSON,
then generate a consolidated verification Excel file per year.

All power values are normalized to KWh on extraction:
  - Source "MW" (or MWh) months are multiplied by 1000.
  - power_unit in JSON is always "KWh" after normalization.

Composite sampling columns are extracted alongside grab columns
wherever they exist (2022+ sheets).  Fields are null when absent.
"""

import json
import os
import re
from datetime import datetime

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_START_ROW = 9
CONTROL_LIMIT_ROW = 7

# ── Per-year column configurations ───────────────────────────────────────────

YEAR_CONFIGS = {
    2022: {
        "months": {
            "January":   "Lab Report January-2022.xlsx",
            "February":  "Lab Report February-2022.xlsx",
            "March":     "Lab Report March-2022.xlsx",
            "April":     "Lab Report April-2022.xlsx",
            "May":       "LAB REPORT may 2022.xlsx",
            "June":      "Lab Report June-2022.xlsx",
            "July":      "Lab Report -July  2022.xlsx",
            "August":    "Lab Report -August 2022.xlsx",
            "September": "Lab Report -Septmeber  2022.xlsx",
            "October":   "Lab Report -October 2022.xlsx",
            "November":  "Lab Report -November 2022.xlsx",
            "December":  "Lab Report -Decemeber  2022 .xlsx",
        },
    },
    2023: {
        "months": {
            "January":   "Lab Report -January 2023.xlsx",
            "February":  "Lab Report -February  2023.xlsx",
            "March":     "Lab Report -March 2023.xlsx",
            "April":     "Lab Report -April month 2023.xlsx",
            "May":       "LAB REPORT MAY 2023.xlsx",
            "June":      "Lab Report June  2023.xlsx",
            "July":      "LAB REPORT JULY 2023.xlsx",
            "August":    "LAB REPORT AUGUST 2023.xlsx",
            "September": "LAB REPORT September 2023.xlsx",
            "October":   "LAB REPORT October 2023.xlsx",
            "November":  "LAB REPORT November 2023.xlsx",
            "December":  "LAB REPORT December 2023.xlsx",
        },
    },
    2024: {
        "months": {
            "January":   "LAB REPORT January 2024.xlsx",
            "February":  "LAB REPORT February 2024.xlsx",
            "March":     "LAB REPORT March 2024.xlsx",
            "April":     "Lab report for April 2024.xlsx",
            "May":       "LAB REPORT May 2024.xlsx",
            "June":      "LAB REPORT June 2024.xlsx",
            "July":      "LAB REPORT July 2024.xlsx",
            "August":    "LAB REPORT August 2024.xlsx",
            "September": "LAB REPORT September 2024.xlsx",
            "October":   "LAB REPORT October 2024.xlsx",
            "November":  "LAB REPORT November 2024.xlsx",
            "December":  "LAB REPORT December 2024.xlsx",
        },
    },
    2025: {
        "months": {
            "January":   "LAB REPORT January 2025.xlsx",
            "February":  "LAB REPORT February 2025.xlsx",
            "March":     "LAB REPORT March 2025.xlsx",
            "April":     "LAB REPORT April 2025.xlsx",
            "May":       "LAB REPORT MAY 2025.xlsx",
            "June":      "LAB REPORT JUNE 2025.xlsx",
            "July":      "LAB REPORT JULY 2025.xlsx",
            "August":    "LAB REPORT AUGUST 2025.xlsx",
            "September": "LAB REPORT SEPT 2025.xlsx",
            "October":   "LAB REPORT OCT 2025.xlsx",
            "November":  "LAB REPORT NOV 2025.xlsx",
            "December":  "LAB REPORT DEC 2025.xlsx",
        },
    },
}

# Keys normalised from MW → KWh
POWER_KEYS = ("power_ge", "power_nea", "power_total")


def detect_columns(ws):
    """
    Auto-detect all column positions from sheet headers.

    Grab (non-composite) columns:
      inlet_ph/bod/cod/tss   — "Raw Sewage" (no COMPOSITE, no FLOW) in row 5
      effluent_ph/bod/cod/tss — "CHLORINE CONTACT TANK" (no COMPOSITE, OUTLET, FLOW) in row 4

    Composite columns (null when absent):
      inlet_*_comp   — "Raw Sewage" + COMPOSITE (no FLOW) in row 5
      effluent_*_comp — "CHLORINE" + COMPOSITE in row 4

    Intermediate stage columns (row 4 section headers):
      grit_tss                — "GRIT" section
      primary_ph/tss/bod/cod/sludge_totalizer — "PRIMARY CLARIFIER" section
      secondary_ph/tss/bod/cod/ras            — "SECONDARY CLARIFIER" section
      sec_sed_ph/tss/bod/cod/ras_new          — "SECONDARY SEDIMENTATION" section

    Extended inlet fields (between inlet and grit sections, row 6 scan):
      inlet_tkn, inlet_og, inlet_po4, inlet_total_coliform, inlet_fecal_coliform

    Extended effluent fields (in CCT region, row 6 scan):
      effluent_og, effluent_nh3, effluent_total_coliform, effluent_fecal_coliform

    Power unit from row 8 col 2.

    Returns (data_cols, ctrl_cols, power_unit).
    """
    r4 = list(ws.iter_rows(min_row=4, max_row=4, values_only=True))[0]
    r5 = list(ws.iter_rows(min_row=5, max_row=5, values_only=True))[0]
    r6 = list(ws.iter_rows(min_row=6, max_row=6, values_only=True))[0]
    r8 = list(ws.iter_rows(min_row=8, max_row=8, values_only=True))[0]

    # ── Power unit ─────────────────────────────────────────────────────────
    power_unit = "KWH" if (r8[1] and "KWH" in str(r8[1]).upper()) else "MW"

    def find_ph_bod_cod_tss(start_col, search_range=20):
        """Scan row 6 from start_col for pH/BOD/COD/TSS. Returns 4-tuple (1-based cols)."""
        ph = bod = cod = tss = None
        for i in range(start_col - 1, min(start_col - 1 + search_range, len(r6))):
            v = r6[i]
            if v is None:
                continue
            s = str(v).upper()
            if "PH" in s and ph is None:
                ph = i + 1
            elif "BOD" in s and bod is None:
                bod = i + 1
            elif "COD" in s and cod is None:
                cod = i + 1
            elif "TSS" in s and tss is None:
                tss = i + 1
            if ph and bod and cod and tss:
                break
        return ph, bod, cod, tss

    def find_section_col(row, must_contain, must_not_contain=()):
        """Return 1-based col of first cell whose text satisfies the criteria."""
        for i, v in enumerate(row):
            if v is None:
                continue
            s = str(v).upper()
            if all(kw in s for kw in must_contain) and not any(kw in s for kw in must_not_contain):
                return i + 1
        return None

    # ── Grab inlet (row 5) ─────────────────────────────────────────────────
    grab_inlet_col = find_section_col(r5, ("RAW SEWAGE",), ("COMPOSITE", "FLOW"))
    inlet_ph = inlet_bod = inlet_cod = inlet_tss = None
    if grab_inlet_col:
        inlet_ph, inlet_bod, inlet_cod, inlet_tss = find_ph_bod_cod_tss(grab_inlet_col)

    # ── Composite inlet (row 5) ────────────────────────────────────────────
    comp_inlet_col = find_section_col(r5, ("RAW SEWAGE", "COMPOSITE"), ("FLOW",))
    inlet_ph_comp = inlet_bod_comp = inlet_cod_comp = inlet_tss_comp = None
    if comp_inlet_col:
        inlet_ph_comp, inlet_bod_comp, inlet_cod_comp, inlet_tss_comp = find_ph_bod_cod_tss(comp_inlet_col)

    # ── Grab CCT effluent (row 4) ──────────────────────────────────────────
    grab_cct_col = find_section_col(r4, ("CHLORINE",), ("COMPOSITE", "OUTLET", "FLOW"))
    eff_ph = eff_bod = eff_cod = eff_tss = eff_frc = None
    if grab_cct_col:
        found = []
        for i in range(grab_cct_col - 1, min(grab_cct_col - 1 + 20, len(r6))):
            v = r6[i]
            if v is None:
                continue
            s = str(v).upper()
            if "PH" in s and eff_ph is None:
                eff_ph = i + 1
                found.append("ph")
            elif "BOD" in s and eff_bod is None:
                eff_bod = i + 1
                found.append("bod")
            elif "COD" in s and eff_cod is None:
                eff_cod = i + 1
                found.append("cod")
            elif "TSS" in s and eff_tss is None:
                eff_tss = i + 1
                found.append("tss")
            elif "FRC" in s and eff_frc is None:
                eff_frc = i + 1
                found.append("frc")
            if len(found) == 5:
                break

    # ── Composite CCT effluent (row 4) ─────────────────────────────────────
    comp_cct_col = find_section_col(r4, ("CHLORINE", "COMPOSITE"))
    eff_ph_comp = eff_bod_comp = eff_cod_comp = eff_tss_comp = None
    if comp_cct_col:
        eff_ph_comp, eff_bod_comp, eff_cod_comp, eff_tss_comp = find_ph_bod_cod_tss(comp_cct_col)

    # ── Grit Classifier (row 4) ────────────────────────────────────────────
    grit_col = find_section_col(r4, ("GRIT",))
    grit_tss = grit_col  # TSS is at the header column itself in row 6

    # ── Primary Clarifier (row 4) ──────────────────────────────────────────
    prim_col = find_section_col(r4, ("PRIMARY", "CLARIFIER"))
    primary_ph = primary_tss = primary_bod = primary_cod = primary_sludge_totalizer = None
    if prim_col:
        primary_ph, primary_bod, primary_cod, primary_tss = find_ph_bod_cod_tss(prim_col)
        if primary_cod:
            for i in range(primary_cod, min(primary_cod + 3, len(r6))):
                v = r6[i]
                if v is not None and ("SLUDGE" in str(v).upper() or "TOTALIZER" in str(v).upper()):
                    primary_sludge_totalizer = i + 1
                    break

    # ── Secondary Clarifier (row 4) ───────────────────────────────────────
    sec_col = find_section_col(r4, ("SECONDARY", "CLARIFIER"), ("SEDIMENTATION",))
    secondary_ph = secondary_tss = secondary_bod = secondary_cod = secondary_ras = None
    if sec_col:
        secondary_ph, secondary_bod, secondary_cod, secondary_tss = find_ph_bod_cod_tss(sec_col)
        if secondary_cod:
            secondary_ras = secondary_cod + 1

    # ── Secondary Sedimentation (row 4) ───────────────────────────────────
    secsed_col = find_section_col(r4, ("SECONDARY", "SEDIMENTATION"))
    sec_sed_ph = sec_sed_tss = sec_sed_bod = sec_sed_cod = sec_sed_ras_new = None
    if secsed_col:
        sec_sed_ph, sec_sed_bod, sec_sed_cod, sec_sed_tss = find_ph_bod_cod_tss(secsed_col)
        if sec_sed_cod:
            sec_sed_ras_new = sec_sed_cod + 1

    # ── Extended inlet fields (scan between inlet sections and grit) ───────
    inlet_tkn = inlet_og = inlet_po4 = inlet_total_coliform = inlet_fecal_coliform = None
    inlet_cols = [x for x in [grab_inlet_col, comp_inlet_col] if x]
    if inlet_cols and grit_col:
        scan_start = min(inlet_cols) + 4
        for i in range(scan_start - 1, grit_col - 1):
            if i >= len(r6):
                break
            v = r6[i]
            if v is None:
                continue
            s = str(v).upper()
            if ("TKN" in s or "NH3" in s) and inlet_tkn is None:
                inlet_tkn = i + 1
            elif ("O & G" in s or ("OIL" in s and "GREAS" in s)) and inlet_og is None:
                inlet_og = i + 1
            elif ("PO4" in s or s.strip() == "TP" or "PHOSPH" in s) and inlet_po4 is None:
                inlet_po4 = i + 1
            elif "TOTAL COLIFORM" in s and inlet_total_coliform is None:
                inlet_total_coliform = i + 1
            elif "FECAL" in s and inlet_fecal_coliform is None:
                inlet_fecal_coliform = i + 1

    # ── Extended effluent fields (scan CCT region) ─────────────────────────
    effluent_og = effluent_nh3 = effluent_total_coliform = effluent_fecal_coliform = None
    cct_cols = [x for x in [grab_cct_col, comp_cct_col] if x]
    if cct_cols:
        cct_scan_start = min(cct_cols)
        for i in range(cct_scan_start - 1, min(cct_scan_start - 1 + 25, len(r6))):
            v = r6[i]
            if v is None:
                continue
            s = str(v).upper()
            if ("O & G" in s or ("OIL" in s and "GREAS" in s)) and effluent_og is None:
                effluent_og = i + 1
            elif ("AMMONI" in s or ("NH3" in s and "INLET" not in s)) and effluent_nh3 is None:
                effluent_nh3 = i + 1
            elif "TOTAL COLIFORM" in s and effluent_total_coliform is None:
                effluent_total_coliform = i + 1
            elif "FECAL" in s and effluent_fecal_coliform is None:
                effluent_fecal_coliform = i + 1

    data_cols = {
        "power_ge":                   2,
        "power_nea":                  3,
        "power_total":                4,
        "power_per_flow":             5,
        "flow":                       6,
        "inlet_ph":                   inlet_ph,
        "inlet_bod":                  inlet_bod,
        "inlet_cod":                  inlet_cod,
        "inlet_tss":                  inlet_tss,
        "inlet_tkn":                  inlet_tkn,
        "inlet_og":                   inlet_og,
        "inlet_po4":                  inlet_po4,
        "inlet_total_coliform":       inlet_total_coliform,
        "inlet_fecal_coliform":       inlet_fecal_coliform,
        "inlet_ph_comp":              inlet_ph_comp,
        "inlet_bod_comp":             inlet_bod_comp,
        "inlet_cod_comp":             inlet_cod_comp,
        "inlet_tss_comp":             inlet_tss_comp,
        "grit_tss":                   grit_tss,
        "primary_ph":                 primary_ph,
        "primary_tss":                primary_tss,
        "primary_bod":                primary_bod,
        "primary_cod":                primary_cod,
        "primary_sludge_totalizer":   primary_sludge_totalizer,
        "secondary_ph":               secondary_ph,
        "secondary_tss":              secondary_tss,
        "secondary_bod":              secondary_bod,
        "secondary_cod":              secondary_cod,
        "secondary_ras":              secondary_ras,
        "sec_sed_ph":                 sec_sed_ph,
        "sec_sed_tss":                sec_sed_tss,
        "sec_sed_bod":                sec_sed_bod,
        "sec_sed_cod":                sec_sed_cod,
        "sec_sed_ras_new":            sec_sed_ras_new,
        "effluent_ph":                eff_ph,
        "effluent_bod":               eff_bod,
        "effluent_cod":               eff_cod,
        "effluent_tss":               eff_tss,
        "effluent_frc":               eff_frc,
        "effluent_og":                effluent_og,
        "effluent_nh3":               effluent_nh3,
        "effluent_total_coliform":    effluent_total_coliform,
        "effluent_fecal_coliform":    effluent_fecal_coliform,
        "effluent_ph_comp":           eff_ph_comp,
        "effluent_bod_comp":          eff_bod_comp,
        "effluent_cod_comp":          eff_cod_comp,
        "effluent_tss_comp":          eff_tss_comp,
    }

    ctrl_cols = {
        "flow":          6,
        "inlet_ph":      inlet_ph,
        "inlet_bod":     inlet_bod,
        "inlet_cod":     inlet_cod,
        "inlet_tss":     inlet_tss,
        "effluent_ph":   eff_ph,
        "effluent_bod":  eff_bod,
        "effluent_cod":  eff_cod,
        "effluent_tss":  eff_tss,
    }

    return data_cols, ctrl_cols, power_unit


# ── Value parsing ─────────────────────────────────────────────────────────────

def parse_value(raw):
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw.strftime("%Y-%m-%d")
    if isinstance(raw, (int, float)):
        return raw
    s = str(raw).strip()
    if s in ("_", "-", "", "BDL", "*BDL", "N/A", "n/a",
             "#DIV/0!", "#REF!", "#N/A", "#VALUE!", "#NAME?", "Nil"):
        return None
    m = re.match(r"^([0-9.]+)\s*[×xX]\s*10\^?(\d+)$", s)
    if m:
        return float(m.group(1)) * (10 ** int(m.group(2)))
    m = re.match(r"^([0-9.]+)\s*[xX]\s*(\d+)$", s)
    if m:
        return float(m.group(1)) * (10 ** int(m.group(2)))
    try:
        return float(s)
    except ValueError:
        return None


def parse_ctrl(raw):
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return raw
    s = str(raw).strip()
    return None if s in ("_", "-", "", "N/A") else s


# ── JSON extraction ───────────────────────────────────────────────────────────

def extract_month(year, month_name, filename):
    year_dir = os.path.join(BASE_DIR, str(year))
    filepath = os.path.join(year_dir, filename)
    if not os.path.exists(filepath):
        print(f"    WARNING: {filename} not found – skipping.")
        return None

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]

    data_cols, ctrl_cols, raw_power_unit = detect_columns(ws)

    # Validate detection of required grab columns
    required = ["inlet_ph", "inlet_bod", "inlet_cod", "inlet_tss",
                "effluent_ph", "effluent_bod", "effluent_cod", "effluent_tss"]
    missing = [k for k in required if data_cols.get(k) is None]
    if missing:
        print(f"    WARNING {month_name}: could not detect grab columns for {missing}")

    # Report composite detection
    comp_keys = ["inlet_ph_comp", "inlet_bod_comp", "inlet_cod_comp", "inlet_tss_comp",
                 "effluent_ph_comp", "effluent_bod_comp", "effluent_cod_comp", "effluent_tss_comp"]
    has_composite = any(data_cols.get(k) is not None for k in comp_keys)

    # Control limits
    control_limits = {"power_per_flow": "< 482.02"}
    for key, col in ctrl_cols.items():
        if col is None:
            continue
        raw = ws.cell(row=CONTROL_LIMIT_ROW, column=col).value
        v = parse_ctrl(raw)
        if v is not None:
            control_limits[key] = v

    # Daily data
    days = []
    row_idx = DATA_START_ROW
    while row_idx < DATA_START_ROW + 50:
        date_cell = ws.cell(row=row_idx, column=1).value
        row_idx += 1
        if date_cell is None:
            break
        if isinstance(date_cell, str):
            s = date_cell.strip().upper()
            if s in ("AVG", "REMARKS:", ""):
                break
            if s == "AVERAGE":
                continue
            break
        if not isinstance(date_cell, datetime):
            break

        day = {"date": date_cell.strftime("%Y-%m-%d")}
        for key, col in data_cols.items():
            day[key] = parse_value(ws.cell(row=row_idx - 1, column=col).value) if col else None

        # Normalize power to KWh (multiply by 1000 if source is MW)
        if raw_power_unit == "MW":
            for pk in POWER_KEYS:
                if day.get(pk) is not None:
                    day[pk] = round(day[pk] * 1000, 3)

        days.append(day)

    wb.close()

    comp_tag = " [+composite]" if has_composite else ""
    return {
        "month": month_name,
        "year": year,
        "power_unit": "KWh",   # always KWh after normalization
        "has_composite": has_composite,
        "control_limits": control_limits,
        "days": days,
        # Internal: kept in memory for verification, stripped before writing to disk
        "_data_cols": data_cols,
        "_comp_tag": comp_tag,
    }


def run_extraction(year):
    cfg = YEAR_CONFIGS[year]
    year_dir = os.path.join(BASE_DIR, str(year))
    data_dir = os.path.join(year_dir, "data")
    os.makedirs(data_dir, exist_ok=True)

    print(f"\n  Extracting {year}...")
    all_months = {}
    for month_name, filename in cfg["months"].items():
        data = extract_month(year, month_name, filename)
        if data is None:
            continue
        mk = month_name.lower()

        # Strip internal fields before saving to disk
        data_to_save = {k: v for k, v in data.items() if not k.startswith("_")}
        out_path = os.path.join(data_dir, f"{mk}.json")
        with open(out_path, "w") as f:
            json.dump(data_to_save, f, indent=2)

        comp_tag = data.get("_comp_tag", "")
        print(f"    {month_name}: {len(data['days'])} days → {mk}.json{comp_tag}")
        all_months[mk] = data  # keep full dict (with _data_cols) for verification

    combined = os.path.join(data_dir, "all_months.json")
    combined_to_save = {
        mk: {k: v for k, v in d.items() if not k.startswith("_")}
        for mk, d in all_months.items()
    }
    with open(combined, "w") as f:
        json.dump(combined_to_save, f, indent=2)
    print(f"    Combined → all_months.json ({len(all_months)} months)")
    return all_months


# ── Excel generation ──────────────────────────────────────────────────────────

DISPLAY_COLS = [
    ("Date",                                         "date"),
    ("Power Gen. from Gas Engine (KWh)",             "power_ge"),
    ("Daily Power Consumption from NEA (KWh)",       "power_nea"),
    ("Total Power Consumption NEA+GE (KWh)",         "power_total"),
    ("Total Power Flow (KWh/ML)",                    "power_per_flow"),
    ("Inlet: Raw Sewage Flow (MLD)",                 "flow"),
    ("Inlet: pH",                                    "inlet_ph"),
    ("Inlet: BOD (mg/L)",                            "inlet_bod"),
    ("Inlet: COD (mg/L)",                            "inlet_cod"),
    ("Inlet: TSS (mg/L)",                            "inlet_tss"),
    ("Effluent: pH",                                 "effluent_ph"),
    ("Effluent: BOD (mg/L)",                         "effluent_bod"),
    ("Effluent: COD (mg/L)",                         "effluent_cod"),
    ("Effluent: TSS (mg/L)",                         "effluent_tss"),
]

CTRL_KEYS = {
    "flow":          "flow",
    "inlet_ph":      "inlet_ph",
    "inlet_bod":     "inlet_bod",
    "inlet_cod":     "inlet_cod",
    "inlet_tss":     "inlet_tss",
    "effluent_ph":   "effluent_ph",
    "effluent_bod":  "effluent_bod",
    "effluent_cod":  "effluent_cod",
    "effluent_tss":  "effluent_tss",
    "power_per_flow":"power_per_flow",
}

def _side(s="thin"):
    return Side(style=s)

def thin_border():
    s = _side()
    return Border(left=s, right=s, top=s, bottom=s)

HDR_MAIN     = PatternFill("solid", fgColor="1F4E79")
HDR_POWER    = PatternFill("solid", fgColor="2E75B6")
HDR_INLET    = PatternFill("solid", fgColor="375623")
HDR_EFFL     = PatternFill("solid", fgColor="833C00")
CTRL_FILL    = PatternFill("solid", fgColor="FFF2CC")
MONTH_FILL   = PatternFill("solid", fgColor="D6E4F0")

WHITE_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
NORMAL_FONT = Font(name="Calibri", size=10)
CTRL_FONT   = Font(name="Calibri", bold=True, color="7F0000", size=9, italic=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left", vertical="center")

def col_fill(jkey):
    if jkey in ("power_ge", "power_nea", "power_total", "power_per_flow"):
        return HDR_POWER
    if jkey.startswith("inlet") or jkey == "flow":
        return HDR_INLET
    if jkey.startswith("effluent"):
        return HDR_EFFL
    return HDR_MAIN

def style_header(cell, jkey):
    cell.fill = col_fill(jkey)
    cell.font = WHITE_FONT
    cell.alignment = CENTER
    cell.border = thin_border()

def style_data(cell, is_date=False):
    cell.font = NORMAL_FONT
    cell.alignment = LEFT if is_date else CENTER
    cell.border = thin_border()

def style_ctrl(cell):
    cell.fill = CTRL_FILL
    cell.font = CTRL_FONT
    cell.alignment = CENTER
    cell.border = thin_border()

def write_header_row(ws, row, col_offset=1):
    for ci, (label, jkey) in enumerate(DISPLAY_COLS):
        cell = ws.cell(row=row, column=col_offset + ci, value=label)
        style_header(cell, jkey)
    return row + 1

def write_ctrl_row(ws, row, control_limits, col_offset=1):
    lbl = ws.cell(row=row, column=col_offset, value="Control Limit")
    lbl.fill = CTRL_FILL; lbl.font = CTRL_FONT
    lbl.alignment = CENTER; lbl.border = thin_border()
    for ci, (_, jkey) in enumerate(DISPLAY_COLS[1:], start=1):
        ck = CTRL_KEYS.get(jkey)
        v = control_limits.get(ck, "—") if ck else "—"
        cell = ws.cell(row=row, column=col_offset + ci, value=v if v is not None else "—")
        style_ctrl(cell)
    return row + 1

def write_data_rows(ws, days, row, col_offset=1):
    for day in days:
        for ci, (_, jkey) in enumerate(DISPLAY_COLS):
            raw = day.get(jkey)
            if jkey == "date" and raw:
                parts = raw.split("-")
                from datetime import date as dtdate
                value = dtdate(int(parts[0]), int(parts[1]), int(parts[2]))
            elif raw is None:
                value = ""
            else:
                value = round(raw, 2) if isinstance(raw, float) else raw
            cell = ws.cell(row=row, column=col_offset + ci, value=value)
            style_data(cell, is_date=(jkey == "date"))
            if jkey == "date" and not isinstance(value, str):
                cell.number_format = "DD-MMM-YYYY"
        row += 1
    return row

def set_col_widths(ws, col_offset=1):
    widths = [14, 14, 14, 14, 14, 12, 8, 10, 10, 10, 8, 10, 10, 10]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(col_offset + i)].width = w

def write_month_banner(ws, label, row, num_cols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    cell = ws.cell(row=row, column=1, value=label)
    cell.fill = MONTH_FILL
    cell.font = Font(name="Calibri", bold=True, size=11, color="1F4E79")
    cell.alignment = CENTER
    cell.border = thin_border()
    return row + 1

def generate_excel(year, all_months_data):
    year_dir = os.path.join(BASE_DIR, str(year))
    output_path = os.path.join(year_dir, f"{year}_Lab_Report_Extracted_Data.xlsx")

    months_order = [
        "january", "february", "march", "april", "may", "june",
        "july", "august", "september", "october", "november", "december"
    ]
    months_present = [m for m in months_order if m in all_months_data]
    month_labels = {k: k.capitalize() for k in months_present}

    num_cols = len(DISPLAY_COLS)
    first_month = months_present[0].capitalize()
    last_month  = months_present[-1].capitalize()

    wb = Workbook()
    wb.remove(wb.active)

    # All Months sheet
    ws_all = wb.create_sheet(f"All Months ({year})")
    ws_all.freeze_panes = "B3"

    ws_all.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    tc = ws_all.cell(row=1, column=1,
        value=f"{year} Lab Report – Extracted Data ({first_month}–{last_month})")
    tc.font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    tc.fill = HDR_MAIN
    tc.alignment = CENTER
    ws_all.row_dimensions[1].height = 22

    current_row = 2
    for mk in months_present:
        data = all_months_data[mk]
        current_row = write_month_banner(ws_all, month_labels[mk], current_row, num_cols)
        current_row = write_header_row(ws_all, current_row)
        current_row = write_ctrl_row(ws_all, current_row, data["control_limits"])
        current_row = write_data_rows(ws_all, data["days"], current_row)
        current_row += 1

    set_col_widths(ws_all)

    # Per-month sheets
    for mk in months_present:
        data = all_months_data[mk]
        ws = wb.create_sheet(month_labels[mk])
        ws.freeze_panes = "B4"

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
        tc = ws.cell(row=1, column=1,
            value=f"{year} Lab Report – {month_labels[mk]} (Extracted Data)")
        tc.font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
        tc.fill = HDR_MAIN
        tc.alignment = CENTER
        ws.row_dimensions[1].height = 22

        row = 2
        row = write_header_row(ws, row)
        row = write_ctrl_row(ws, row, data["control_limits"])
        write_data_rows(ws, data["days"], row)
        set_col_widths(ws)

    wb.save(output_path)
    print(f"    Excel saved: {os.path.basename(output_path)}")
    return output_path


# ── Verification ──────────────────────────────────────────────────────────────

def verify_chain(year, all_months_data):
    """Verify JSON grab values against source Excel using auto-detected columns."""
    cfg = YEAR_CONFIGS[year]
    year_dir = os.path.join(BASE_DIR, str(year))

    total_days = 0
    total_errors = 0

    for month_name, filename in cfg["months"].items():
        mk = month_name.lower()
        if mk not in all_months_data:
            continue

        filepath = os.path.join(year_dir, filename)
        if not os.path.exists(filepath):
            continue

        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        ws = wb.active

        data_cols = all_months_data[mk].get("_data_cols", {})
        if not data_cols:
            data_cols, _, _ = detect_columns(ws)

        json_days = {d["date"]: d for d in all_months_data[mk]["days"]}

        errors = []
        matched = 0

        for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
            date_val = row[0]
            if not isinstance(date_val, datetime):
                if isinstance(date_val, str) and date_val.strip().upper() == "AVERAGE":
                    continue
                continue
            date_str = date_val.strftime("%Y-%m-%d")
            if date_str not in json_days:
                errors.append(f"  {date_str}: in Excel but missing in JSON")
                continue
            matched += 1
            day = json_days[date_str]

            # Determine if this sheet uses MW (need to scale for comparison)
            raw_power_unit = "MW"
            r8 = list(ws.iter_rows(min_row=8, max_row=8, values_only=True))[0]
            if r8[1] and "KWH" in str(r8[1]).upper():
                raw_power_unit = "KWH"

            for key, col in data_cols.items():
                if col is None:
                    continue
                xl_raw = row[col - 1]
                xl_val = parse_value(xl_raw)
                json_val = day.get(key)

                # Scale power for comparison if source was MW
                if key in POWER_KEYS and raw_power_unit == "MW" and xl_val is not None:
                    xl_val = xl_val * 1000

                def norm(v):
                    if v is None:
                        return None
                    return round(v, 1) if isinstance(v, float) else v

                xn, jn = norm(xl_val), norm(json_val)
                if xn != jn:
                    if isinstance(xn, float) and isinstance(jn, float) and abs(xn - jn) < 0.5:
                        continue
                    errors.append(f"  {date_str} [{key}]: XL={xn!r}  JSON={jn!r}")

        wb.close()
        total_days += matched
        total_errors += len(errors)

        if errors:
            print(f"    {month_name}: {matched} days, {len(errors)} discrepancies:")
            for e in errors[:5]:
                print(f"      {e}")
            if len(errors) > 5:
                print(f"      ... and {len(errors)-5} more")
        else:
            print(f"    {month_name}: {matched} days — OK")

    print(f"  → {year} total: {total_days} days, {total_errors} discrepancies")
    return total_errors


# ── Main ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    grand_total_errors = 0

    for year in [2022, 2023, 2024, 2025]:
        print(f"\n{'='*60}")
        print(f"  YEAR {year}")
        print(f"{'='*60}")

        all_months = run_extraction(year)
        generate_excel(year, all_months)

        print(f"\n  Verifying chain (source Excel → JSON)...")
        errs = verify_chain(year, all_months)
        grand_total_errors += errs

    print(f"\n{'='*60}")
    print(f"ALL DONE – grand total discrepancies: {grand_total_errors}")
