"""
Merge extracted lab report data from 2020 CSV + 2021–2025 JSON into a single Excel file.

Output: 21-25/All_Years_Merged.xlsx
  - One header row (no month/year banners between rows)
  - Data ordered chronologically (2020 Oct → 2025 Dec)
  - All power values in KW (CSV MWh values are multiplied ×1000)
  - JSON data takes priority over CSV for any overlapping dates
"""

import json
import os
from datetime import date as dtdate

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# ── Column schema ──────────────────────────────────────────────────────────────
# Only columns present in ALL years (2021 Jul–Dec JSON + 2022–2025 JSON).
# Composite fields are null for 2021 (has_composite=False) and often null for
# individual days in 2022-2025 — they are included so the merged file stays
# useful for analysis.
COLUMNS = [
    ("Date",                                   "date"),
    ("Power GE (KW)",                          "power_ge"),
    ("Power NEA (KW)",                         "power_nea"),
    ("Power Total (KW)",                       "power_total"),
    ("Power / Flow (KW/ML)",                   "power_per_flow"),
    ("Flow (MLD)",                             "flow"),
    ("Inlet pH (Grab)",                        "inlet_ph"),
    ("Inlet BOD (mg/L, Grab)",                 "inlet_bod"),
    ("Inlet COD (mg/L, Grab)",                 "inlet_cod"),
    ("Inlet TSS (mg/L, Grab)",                 "inlet_tss"),
    ("Inlet pH (Composite)",                   "inlet_ph_comp"),
    ("Inlet BOD (mg/L, Composite)",            "inlet_bod_comp"),
    ("Inlet COD (mg/L, Composite)",            "inlet_cod_comp"),
    ("Inlet TSS (mg/L, Composite)",            "inlet_tss_comp"),
    ("Primary TSS (mg/L)",                     "primary_tss"),
    ("Primary BOD (mg/L)",                     "primary_bod"),
    ("Primary COD (mg/L)",                     "primary_cod"),
    ("Sec Clarifier pH",                       "secondary_ph"),
    ("Sec Clarifier TSS (mg/L)",               "secondary_tss"),
    ("Sec Clarifier BOD (mg/L)",               "secondary_bod"),
    ("Sec Clarifier COD (mg/L)",               "secondary_cod"),
    ("Sec Clarifier RAS",                      "secondary_ras"),
    ("Sec Sed pH",                             "sec_sed_ph"),
    ("Sec Sed TSS (mg/L)",                     "sec_sed_tss"),
    ("Sec Sed BOD (mg/L)",                     "sec_sed_bod"),
    ("Sec Sed COD (mg/L)",                     "sec_sed_cod"),
    ("Sec Sed RAS (New)",                      "sec_sed_ras_new"),
    ("Effluent pH (Grab)",                     "effluent_ph"),
    ("Effluent BOD (mg/L, Grab)",              "effluent_bod"),
    ("Effluent COD (mg/L, Grab)",              "effluent_cod"),
    ("Effluent TSS (mg/L, Grab)",              "effluent_tss"),
    ("Effluent FRC (mg/L)",                    "effluent_frc"),
    ("Effluent pH (Composite)",                "effluent_ph_comp"),
    ("Effluent BOD (mg/L, Composite)",         "effluent_bod_comp"),
    ("Effluent COD (mg/L, Composite)",         "effluent_cod_comp"),
    ("Effluent TSS (mg/L, Composite)",         "effluent_tss_comp"),
]

MONTH_ORDER = [
    "january", "february", "march", "april", "may", "june",
    "july", "august", "september", "october", "november", "december",
]

# ── Styles ─────────────────────────────────────────────────────────────────────
def _side(s="thin"):
    return Side(style=s)

def thin_border():
    s = _side()
    return Border(left=s, right=s, top=s, bottom=s)

HDR_MAIN      = PatternFill("solid", fgColor="1F4E79")
HDR_POWER     = PatternFill("solid", fgColor="2E75B6")
HDR_INLET     = PatternFill("solid", fgColor="375623")
HDR_EFFL      = PatternFill("solid", fgColor="833C00")
HDR_PRIMARY   = PatternFill("solid", fgColor="7030A0")
HDR_SECONDARY = PatternFill("solid", fgColor="006B6B")
HDR_SECSED    = PatternFill("solid", fgColor="2E4057")

WHITE_BOLD  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
NORMAL_FONT = Font(name="Calibri", size=10)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center")

def col_fill(jkey):
    if "power" in jkey:
        return HDR_POWER
    if jkey.startswith("inlet") or jkey == "flow":
        return HDR_INLET
    if jkey.startswith("primary"):
        return HDR_PRIMARY
    if jkey.startswith("secondary"):
        return HDR_SECONDARY
    if jkey.startswith("sec_sed"):
        return HDR_SECSED
    if jkey.startswith("effluent"):
        return HDR_EFFL
    return HDR_MAIN

# ── Data loading ───────────────────────────────────────────────────────────────

def load_json_year(year):
    """Load all months from the year's all_months.json, return list of day dicts."""
    path = os.path.join(BASE_DIR, "raw_data", str(year), "extracted_data", "all_months.json")
    if not os.path.exists(path):
        print(f"  WARNING: {path} not found — skipping {year}.")
        return []
    with open(path) as f:
        all_months = json.load(f)

    rows = []
    for mk in MONTH_ORDER:
        if mk not in all_months:
            continue
        for day in all_months[mk]["days"]:
            rows.append(day)
    return rows


def load_csv(csv_path):
    """
    Load the 2020 RAW.csv and convert to the common schema.

    CSV columns: Date, Q_IN, IN_BOD, IN_COD, IN_TSS, IN_pH, PC, PF, O_BOD, O_COD, O_TSS

    PC is the daily NEA power consumption in MWh → multiply ×1000 to get KW.
    PF (power/flow ratio) is already in KW/ML — used directly.
    """
    import csv
    from datetime import datetime

    rows = []
    with open(csv_path, newline="") as f:
        reader = csv.DictReader(f)
        for rec in reader:
            def v(col):
                raw = rec.get(col, "").strip()
                if raw in ("", "-", "N/A"):
                    return None
                try:
                    return float(raw)
                except ValueError:
                    return None

            date_str = rec.get("Date", "").strip()
            try:
                dt = datetime.strptime(date_str, "%m/%d/%Y")
                date_fmt = dt.strftime("%Y-%m-%d")
            except ValueError:
                date_fmt = date_str

            pc_mwh = v("PC")
            pc_kw  = pc_mwh * 1000 if pc_mwh is not None else None   # MWh → KW

            row = {
                "date":              date_fmt,
                "power_ge":          None,        # GE not tracked in CSV
                "power_nea":         pc_kw,
                "power_total":       pc_kw,       # no GE contribution in CSV period
                "power_per_flow":    v("PF"),     # already in KW/ML
                "flow":              v("Q_IN"),
                "inlet_ph":          v("IN_pH"),
                "inlet_bod":         v("IN_BOD"),
                "inlet_cod":         v("IN_COD"),
                "inlet_tss":         v("IN_TSS"),
                "inlet_ph_comp":     None,
                "inlet_bod_comp":    None,
                "inlet_cod_comp":    None,
                "inlet_tss_comp":    None,
                "effluent_ph":       None,        # not in CSV
                "effluent_bod":      v("O_BOD"),
                "effluent_cod":      v("O_COD"),
                "effluent_tss":      v("O_TSS"),
                "effluent_frc":      None,
                "effluent_ph_comp":  None,
                "effluent_bod_comp": None,
                "effluent_cod_comp": None,
                "effluent_tss_comp": None,
            }
            rows.append(row)
    return rows


def merge_all_rows():
    """
    Merge rows in chronological order: 2020 CSV → 2021–2025 JSON.

    The CSV spans Oct 2020 – Jun 2022 so it overlaps with 2021 Jul–Dec and
    2022 Jan–Jun JSON data. JSON always takes priority for overlapping dates
    since it carries more attributes (power_ge, composite samples, etc.).
    """
    CSV_PATH = os.path.join(BASE_DIR, "raw_data", "2020", "RAW.csv")

    # ── Collect all JSON dates (priority source) ───────────────────────────────
    json_rows = []
    for year in [2021, 2022, 2023, 2024, 2025]:
        rows = load_json_year(year)
        json_rows.extend(rows)
        print(f"  {year}: {len(rows)} rows (JSON)")
    json_dates = {r["date"] for r in json_rows}

    # ── Load CSV, skip any date already covered by JSON ────────────────────────
    csv_rows = load_csv(CSV_PATH)
    csv_kept = [r for r in csv_rows if r["date"] not in json_dates]
    csv_skip = len(csv_rows) - len(csv_kept)
    print(f"  CSV:  {len(csv_rows)} rows total, {csv_skip} skipped (overlap with JSON), "
          f"{len(csv_kept)} added")

    all_rows = csv_kept + json_rows
    all_rows.sort(key=lambda r: r.get("date", ""))
    return all_rows


# ── Excel writing ──────────────────────────────────────────────────────────────

def write_excel(rows, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "All Years 2020–2025"
    ws.freeze_panes = "B2"

    num_cols = len(COLUMNS)

    # Header row
    for ci, (label, jkey) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=ci, value=label)
        cell.fill = col_fill(jkey)
        cell.font = WHITE_BOLD
        cell.alignment = CENTER
        cell.border = thin_border()

    ws.row_dimensions[1].height = 36

    # Data rows
    for ri, day in enumerate(rows, start=2):
        for ci, (_, jkey) in enumerate(COLUMNS, start=1):
            raw = day.get(jkey)
            if jkey == "date" and raw:
                parts = raw.split("-")
                try:
                    value = dtdate(int(parts[0]), int(parts[1]), int(parts[2]))
                    fmt = "DD-MMM-YYYY"
                except Exception:
                    value = raw
                    fmt = None
            elif raw is None:
                value = ""
                fmt = None
            elif isinstance(raw, float):
                value = round(raw, 3)
                fmt = None
            else:
                value = raw
                fmt = None

            cell = ws.cell(row=ri, column=ci, value=value)
            cell.font = NORMAL_FONT
            cell.alignment = LEFT if jkey == "date" else CENTER
            cell.border = thin_border()
            if jkey == "date" and fmt:
                cell.number_format = fmt

    # Column widths
    widths = [
        13, 12, 12, 12, 14, 9,          # Date, Power x4, Flow
        10, 14, 14, 14,                  # Inlet grab (pH, BOD, COD, TSS)
        14, 18, 18, 18,                  # Inlet composite (pH, BOD, COD, TSS)
        14, 14, 14,                      # Primary (TSS, BOD, COD)
        14, 16, 16, 16, 14,              # Sec Clarifier (pH, TSS, BOD, COD, RAS)
        12, 16, 16, 16, 14,              # Sec Sed (pH, TSS, BOD, COD, RAS)
        13, 17, 17, 17, 12,              # Effluent grab (pH, BOD, COD, TSS, FRC)
        17, 21, 21, 21,                  # Effluent composite (pH, BOD, COD, TSS)
    ]
    for i, w in enumerate(widths[:num_cols], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(output_path)
    print(f"\n  Saved: {output_path}")
    print(f"  Total rows: {len(rows)}")


# ── Main ───────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("Merging all years (2021–2025)...")
    rows = merge_all_rows()
    output = os.path.join(BASE_DIR, "All_Years_Merged.xlsx")
    write_excel(rows, output)
    print("Done.")
