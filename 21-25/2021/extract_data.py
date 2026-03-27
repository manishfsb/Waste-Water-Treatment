#!/usr/bin/env python3
"""
Extract data from wastewater treatment lab report Excel files into JSON.
Reads all monthly reports from the 2021 directory and outputs structured JSON to data/.
"""

import json
import os
import re
import sys
from datetime import datetime

import openpyxl

# Column mapping: JSON key -> Excel column letter
# Organized by treatment stage
COLUMN_MAP = {
    # Power
    "power_ge": "B",
    "power_nea": "C",
    "power_total": "D",
    "power_per_flow": "E",
    # Inlet
    "flow": "F",
    "inlet_ph": "G",
    "inlet_bod": "H",
    "inlet_cod": "I",
    "inlet_tss": "J",
    "inlet_tkn": "K",
    "inlet_og": "L",
    "inlet_po4": "M",
    "inlet_total_coliform": "N",
    "inlet_fecal_coliform": "O",
    # Grit Classifier
    "grit_tss": "Q",
    # Primary Clarifier
    "primary_ph": "R",
    "primary_tss": "S",
    "primary_bod": "T",
    "primary_cod": "U",
    "primary_sludge_totalizer": "V",
    # Secondary Clarifier
    "secondary_ph": "W",
    "secondary_tss": "X",
    "secondary_bod": "Y",
    "secondary_cod": "Z",
    "secondary_ras": "AA",
    # Secondary Sedimentation Outlet
    "sec_sed_ph": "AB",
    "sec_sed_tss": "AC",
    "sec_sed_bod": "AD",
    "sec_sed_cod": "AE",
    "sec_sed_ras_new": "AF",
    # Effluent (Chlorine Contact Tank)
    "effluent_ph": "AH",
    "effluent_bod": "AI",
    "effluent_cod": "AJ",
    "effluent_tss": "AK",
    "effluent_og": "AL",
    "effluent_nh3": "AM",
    "effluent_total_coliform": "AN",
    "effluent_fecal_coliform": "AO",
    "effluent_frc": "AP",
}

# Control limit column mapping (row 7) - same column letters as data
CONTROL_LIMIT_KEYS = {
    "flow": "F",
    "inlet_ph": "G",
    "inlet_bod": "H",
    "inlet_cod": "I",
    "inlet_tss": "J",
    "inlet_tkn": "K",
    "inlet_og": "L",
    "inlet_po4": "M",
    "inlet_total_coliform": "N",
    "inlet_fecal_coliform": "O",
    "effluent_ph": "AH",
    "effluent_bod": "AI",
    "effluent_cod": "AJ",
    "effluent_tss": "AK",
    "effluent_og": "AL",
    "effluent_nh3": "AM",
    "effluent_total_coliform": "AN",
    "effluent_fecal_coliform": "AO",
    "effluent_frc": "AP",
    "power_per_flow": "E",
}

MONTH_FILES = {
    "July": "Lab Report JULY-2021.xlsx",
    "August": "Lab Report AUGUST-2021.xlsx",
    "September": "Lab Report September-2021.xlsx",
    "October": "Lab Report October-2021.xlsx",
    "November": "Lab Report November-2021.xlsx",
    "December": "Lab Report December-2021.xlsx",
}

HEADER_ROW = 6
CONTROL_LIMIT_ROW = 7
UNITS_ROW = 8
DATA_START_ROW = 9


def col_letter_to_index(letter):
    """Convert Excel column letter (A, B, ..., AA, AB, ...) to 1-based index."""
    result = 0
    for char in letter.upper():
        result = result * 26 + (ord(char) - ord("A") + 1)
    return result


def parse_value(raw):
    """Parse a raw cell value into a numeric value or None."""
    if raw is None:
        return None

    # Handle datetime objects (dates in column A)
    if isinstance(raw, datetime):
        return raw.isoformat()

    # Handle numeric types directly
    if isinstance(raw, (int, float)):
        return raw

    # String handling
    s = str(raw).strip()

    # Missing value markers
    if s in ("_", "-", "", "BDL", "*BDL", "N/A", "n/a",
             "#DIV/0!", "#REF!", "#N/A", "#VALUE!", "#NAME?"):
        return None

    # Scientific notation with × (e.g., "3.8×106", "5.0×105")
    match = re.match(r"^([0-9.]+)\s*[×xX]\s*10\^?(\d+)$", s)
    if match:
        base = float(match.group(1))
        exp = int(match.group(2))
        return base * (10 ** exp)

    # Also handle formats like "3.5X 104"
    match = re.match(r"^([0-9.]+)\s*[xX]\s*(\d+)$", s)
    if match:
        base = float(match.group(1))
        exp = int(match.group(2))
        return base * (10 ** exp)

    # Try direct float conversion
    try:
        return float(s)
    except ValueError:
        pass

    # Return the string as-is (will be flagged in validation)
    return s


def parse_control_limit(raw):
    """Parse a control limit value - keep as string for display."""
    if raw is None:
        return "N/A"

    if isinstance(raw, (int, float)):
        return raw

    s = str(raw).strip()
    if s in ("_", "-", ""):
        return "N/A"

    return s


def extract_month(base_dir, month_name, filename):
    """Extract data from a single month's Excel file."""
    filepath = os.path.join(base_dir, filename)
    if not os.path.exists(filepath):
        print(f"  WARNING: {filepath} not found, skipping.")
        return None

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # Extract control limits from row 7
    control_limits = {}
    for key, col_letter in CONTROL_LIMIT_KEYS.items():
        col_idx = col_letter_to_index(col_letter)
        raw = ws.cell(row=CONTROL_LIMIT_ROW, column=col_idx).value
        control_limits[key] = parse_control_limit(raw)

    # power_per_flow limit is in the units row (C8) as text, not in row 7
    control_limits["power_per_flow"] = "< 482.02"

    # Extract daily data (rows 9 until we hit 'AVG' or empty date)
    days = []
    for row_idx in range(DATA_START_ROW, DATA_START_ROW + 32):  # max 31 days
        date_cell = ws.cell(row=row_idx, column=1).value
        if date_cell is None or (isinstance(date_cell, str) and date_cell.strip() in ("AVG", "Remarks:", "")):
            break

        day_data = {"date": None}

        # Parse date
        if isinstance(date_cell, datetime):
            day_data["date"] = date_cell.strftime("%Y-%m-%d")
        else:
            day_data["date"] = str(date_cell)

        # Parse all mapped columns
        for key, col_letter in COLUMN_MAP.items():
            col_idx = col_letter_to_index(col_letter)
            raw = ws.cell(row=row_idx, column=col_idx).value
            day_data[key] = parse_value(raw)

        days.append(day_data)

    result = {
        "month": month_name,
        "year": 2021,
        "control_limits": control_limits,
        "days": days,
    }

    return result


def main():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    out_dir = os.path.join(base_dir, "data")
    os.makedirs(out_dir, exist_ok=True)

    print("Extracting wastewater lab report data...")
    print(f"  Source: {base_dir}")
    print(f"  Output: {out_dir}")
    print()

    all_months = {}
    for month_name, filename in MONTH_FILES.items():
        print(f"  Processing {month_name}...")
        data = extract_month(base_dir, month_name, filename)
        if data:
            out_file = os.path.join(out_dir, f"{month_name.lower()}.json")
            with open(out_file, "w") as f:
                json.dump(data, f, indent=2)
            print(f"    -> {out_file} ({len(data['days'])} days)")
            all_months[month_name.lower()] = data

    # Also write a combined file for convenience
    combined_file = os.path.join(out_dir, "all_months.json")
    with open(combined_file, "w") as f:
        json.dump(all_months, f, indent=2)
    print(f"\n  Combined -> {combined_file}")
    print("\nDone!")


if __name__ == "__main__":
    main()
