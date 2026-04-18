"""
Generate Excel file from 2021 lab report JSON data.
Columns: Date, Power (GE, NEA, Total, Power/Flow), Inlet (Flow, pH, BOD, COD, TSS),
         Effluent (pH, BOD, COD, TSS) with control limits.
"""

import json
import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

DATA_DIR = os.path.join(os.path.dirname(__file__), "data")
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), "2021_Lab_Report_Extracted_Data.xlsx")

MONTHS_ORDER = ["july", "august", "september", "october", "november", "december"]
MONTH_LABELS = {
    "july": "July", "august": "August", "september": "September",
    "october": "October", "november": "November", "december": "December"
}

# Columns to extract: (display_label, json_key)
COLUMNS = [
    ("Date",                                    "date"),
    # Power
    ("Power Gen. from Gas Engine (MWh)",        "power_ge"),
    ("Daily Power Consumption from NEA (MWh)",  "power_nea"),
    ("Total Power Consumption NEA+GE (MWh)",    "power_total"),
    ("Total Power Flow (KWh/ML)",               "power_per_flow"),
    # Inlet
    ("Inlet: Raw Sewage Flow (MLD)",            "flow"),
    ("Inlet: pH",                               "inlet_ph"),
    ("Inlet: BOD (mg/L)",                       "inlet_bod"),
    ("Inlet: COD (mg/L)",                       "inlet_cod"),
    ("Inlet: TSS (mg/L)",                       "inlet_tss"),
    # Effluent (Chlorine Contact Tank)
    ("Effluent: pH",                            "effluent_ph"),
    ("Effluent: BOD (mg/L)",                    "effluent_bod"),
    ("Effluent: COD (mg/L)",                    "effluent_cod"),
    ("Effluent: TSS (mg/L)",                    "effluent_tss"),
]

# Control limit keys matching the column json_keys above
CONTROL_LIMIT_KEYS = {
    "flow":          "flow",
    "inlet_ph":      "inlet_ph",
    "inlet_bod":     "inlet_bod",
    "inlet_cod":     "inlet_cod",
    "inlet_tss":     "inlet_tss",
    "effluent_ph":   "effluent_ph",
    "effluent_bod":  "effluent_bod",
    "effluent_cod":  "effluent_cod",
    "effluent_tss":  "effluent_tss",
    "power_per_flow":"power_per_flow",
}

# ── Styles ──────────────────────────────────────────────────────────────────

def _side(style="thin"):
    return Side(style=style)

def thin_border():
    s = _side()
    return Border(left=s, right=s, top=s, bottom=s)

HEADER_FILL_MAIN    = PatternFill("solid", fgColor="1F4E79")   # dark blue
HEADER_FILL_POWER   = PatternFill("solid", fgColor="2E75B6")   # mid blue
HEADER_FILL_INLET   = PatternFill("solid", fgColor="375623")   # dark green
HEADER_FILL_EFFLUENT= PatternFill("solid", fgColor="833C00")   # dark orange
CTRL_FILL           = PatternFill("solid", fgColor="FFF2CC")   # light yellow
MONTH_LABEL_FILL    = PatternFill("solid", fgColor="D6E4F0")   # very light blue

WHITE_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
DARK_FONT   = Font(name="Calibri", bold=True, color="000000", size=10)
NORMAL_FONT = Font(name="Calibri", size=10)
CTRL_FONT   = Font(name="Calibri", bold=True, color="7F0000", size=9, italic=True)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center")


def _col_fill(json_key):
    """Return header fill colour based on which section the column belongs to."""
    if json_key in ("power_ge", "power_nea", "power_total", "power_per_flow"):
        return HEADER_FILL_POWER
    if json_key.startswith("inlet") or json_key == "flow":
        return HEADER_FILL_INLET
    if json_key.startswith("effluent"):
        return HEADER_FILL_EFFLUENT
    return HEADER_FILL_MAIN


def style_header_cell(cell, json_key):
    cell.fill   = _col_fill(json_key)
    cell.font   = WHITE_FONT
    cell.alignment = CENTER
    cell.border = thin_border()


def style_data_cell(cell, is_date=False):
    cell.font      = NORMAL_FONT
    cell.alignment = LEFT if is_date else CENTER
    cell.border    = thin_border()


def style_ctrl_cell(cell):
    cell.fill      = CTRL_FILL
    cell.font      = CTRL_FONT
    cell.alignment = CENTER
    cell.border    = thin_border()


# ── Data loading ─────────────────────────────────────────────────────────────

def load_month(month_key):
    path = os.path.join(DATA_DIR, f"{month_key}.json")
    with open(path) as f:
        return json.load(f)


# ── Sheet writing ─────────────────────────────────────────────────────────────

def write_header_row(ws, row, col_offset=1):
    """Write the column header row and return the row index after it."""
    for ci, (label, jkey) in enumerate(COLUMNS):
        cell = ws.cell(row=row, column=col_offset + ci, value=label)
        style_header_cell(cell, jkey)
    return row + 1


def write_control_limit_row(ws, row, control_limits, col_offset=1):
    """Write a 'Control Limit' row beneath the header."""
    # Label cell under Date column
    lbl = ws.cell(row=row, column=col_offset, value="Control Limit")
    lbl.fill = CTRL_FILL
    lbl.font = CTRL_FONT
    lbl.alignment = CENTER
    lbl.border = thin_border()

    for ci, (_, jkey) in enumerate(COLUMNS[1:], start=1):
        cl_key = CONTROL_LIMIT_KEYS.get(jkey)
        value = control_limits.get(cl_key, "") if cl_key else ""
        cell = ws.cell(row=row, column=col_offset + ci, value=value if value != "" else "—")
        style_ctrl_cell(cell)
    return row + 1


def write_data_rows(ws, days, row, col_offset=1):
    """Write daily data rows. Returns next available row."""
    for day in days:
        for ci, (_, jkey) in enumerate(COLUMNS):
            raw = day.get(jkey)
            if jkey == "date" and raw:
                from datetime import date as dt_date
                parts = raw.split("-")
                value = dt_date(int(parts[0]), int(parts[1]), int(parts[2]))
            elif raw is None:
                value = ""
            else:
                # Round power_per_flow to 2 dp for readability
                value = round(raw, 2) if isinstance(raw, float) else raw
            cell = ws.cell(row=row, column=col_offset + ci, value=value)
            style_data_cell(cell, is_date=(jkey == "date"))
            if jkey == "date" and isinstance(value, __import__("datetime").date):
                cell.number_format = "DD-MMM-YYYY"
        row += 1
    return row


def set_column_widths(ws, col_offset=1):
    widths = [14, 14, 14, 14, 14, 12, 8, 10, 10, 10, 8, 10, 10, 10]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(col_offset + i)].width = w


# ── Section header for merged-all sheet ──────────────────────────────────────

def write_month_label_row(ws, month_label, row, num_cols):
    ws.merge_cells(
        start_row=row, start_column=1,
        end_row=row,   end_column=num_cols
    )
    cell = ws.cell(row=row, column=1, value=month_label)
    cell.fill      = MONTH_LABEL_FILL
    cell.font      = Font(name="Calibri", bold=True, size=11, color="1F4E79")
    cell.alignment = CENTER
    cell.border    = thin_border()
    return row + 1


# ── Build workbook ────────────────────────────────────────────────────────────

def build_workbook():
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    all_month_data = []
    for mk in MONTHS_ORDER:
        data = load_month(mk)
        all_month_data.append((mk, data))

    num_cols = len(COLUMNS)

    # ── 1. "All Months" merged sheet ─────────────────────────────────────────
    ws_all = wb.create_sheet("All Months (2021)")
    ws_all.freeze_panes = "B3"   # freeze after date col + header rows

    # Title
    ws_all.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    title_cell = ws_all.cell(row=1, column=1,
        value="2021 Lab Report – Extracted Data (Jul–Dec)")
    title_cell.font      = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    title_cell.fill      = PatternFill("solid", fgColor="1F4E79")
    title_cell.alignment = CENTER
    ws_all.row_dimensions[1].height = 22

    current_row = 2

    for mk, data in all_month_data:
        month_label = MONTH_LABELS[mk]
        control_limits = data.get("control_limits", {})
        days = data.get("days", [])

        # Month label banner
        current_row = write_month_label_row(ws_all, month_label, current_row, num_cols)
        # Header row
        current_row = write_header_row(ws_all, current_row)
        # Control limit row
        current_row = write_control_limit_row(ws_all, current_row, control_limits)
        # Data rows
        current_row = write_data_rows(ws_all, days, current_row)
        # Blank separator
        current_row += 1

    set_column_widths(ws_all)

    # ── 2. Individual monthly sheets ──────────────────────────────────────────
    for mk, data in all_month_data:
        month_label = MONTH_LABELS[mk]
        control_limits = data.get("control_limits", {})
        days = data.get("days", [])

        ws = wb.create_sheet(month_label)
        ws.freeze_panes = "B4"  # freeze after header + control limit rows

        # Title
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
        tc = ws.cell(row=1, column=1,
            value=f"2021 Lab Report – {month_label} (Extracted Data)")
        tc.font      = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
        tc.fill      = PatternFill("solid", fgColor="1F4E79")
        tc.alignment = CENTER
        ws.row_dimensions[1].height = 22

        row = 2
        row = write_header_row(ws, row)
        row = write_control_limit_row(ws, row, control_limits)
        write_data_rows(ws, days, row)
        set_column_widths(ws)

    wb.save(OUTPUT_PATH)
    print(f"Saved: {OUTPUT_PATH}")
    return OUTPUT_PATH


if __name__ == "__main__":
    build_workbook()
