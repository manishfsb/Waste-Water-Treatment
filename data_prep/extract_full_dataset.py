"""
Extract the complete feature set from all monthly lab report Excel files
(2021 Jul-Dec, 2022-2025) and the 2020 RAW.csv, producing All_Years_Full.xlsx.

This script is standalone - it does NOT modify any existing files
(process_years.py, extract_data.py, merge_all_years.py, JSON files, or
All_Years_Merged.xlsx).

New columns vs All_Years_Merged.xlsx:
  Inlet grab    : TKN (2021 only), O&G (2021 only), PO4/TP (2021-2022), Total/Fecal Coliform (2021-2022)
  Inlet composite: NH3-N (2022+), O&G (2022+), PO4/TP (2023+), Total/Fecal Coliform (2023+)
  Grit    : TSS
  Primary : pH, Sludge Totalizer  (TSS/BOD/COD were already present)
  Aeration: pH, DO, MLSS, MLVSS, SV30, SVI  (x2: existing + new tank)
  Effluent: O&G, NH3-N, Total Coliform, Fecal Coliform

Extended inlet feature provenance:
  TKN/NH3-N and O&G switched from grab to composite measurement in 2022.
  PO4/TP and Coliforms switched in 2023. Each feature is now stored in TWO
  columns: grab-sourced (sparse, early years) and composite-sourced (dense,
  later years). Use the correct column for your target type.

Source priority for overlapping dates: Excel files > 2020 RAW.csv
2021 Jan-Jun has no lab report Excel files - those rows come from the CSV only
(limited columns; all new fields will be NaN for those dates).
"""

import csv
import os
import re
from datetime import date as dtdate
from datetime import datetime

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_START_ROW = 9   # same for all years
POWER_KEYS = ("power_ge", "power_nea", "power_total")

# ── Year → month → filename ────────────────────────────────────────────────────

YEAR_FILES = {
    2021: {
        "July":      "Lab Report JULY-2021.xlsx",
        "August":    "Lab Report AUGUST-2021.xlsx",
        "September": "Lab Report September-2021.xlsx",
        "October":   "Lab Report October-2021.xlsx",
        "November":  "Lab Report November-2021.xlsx",
        "December":  "Lab Report December-2021.xlsx",
    },
    2022: {
        "January":   "Lab Report January-2022.xlsx",
        "February":  "Lab Report February-2022.xlsx",
        "March":     "Lab Report March-2022.xlsx",
        "April":     "Lab Report April-2022.xlsx",
        "May":       "LAB REPORT may 2022.xlsx",
        "June":      "Lab Report June-2022.xlsx",
        "July":      "Lab Report -July  2022.xlsx",
        "August":    "Lab Report -August 2022.xlsx",
        "September": "Lab Report -Septmeber  2022.xlsx",
        "October":   "Lab Report -October 2022.xlsx",
        "November":  "Lab Report -November 2022.xlsx",
        "December":  "Lab Report -Decemeber  2022 .xlsx",
    },
    2023: {
        "January":   "Lab Report -January 2023.xlsx",
        "February":  "Lab Report -February  2023.xlsx",
        "March":     "Lab Report -March 2023.xlsx",
        "April":     "Lab Report -April month 2023.xlsx",
        "May":       "LAB REPORT MAY 2023.xlsx",
        "June":      "Lab Report June  2023.xlsx",
        "July":      "LAB REPORT JULY 2023.xlsx",
        "August":    "LAB REPORT AUGUST 2023.xlsx",
        "September": "LAB REPORT September 2023.xlsx",
        "October":   "LAB REPORT October 2023.xlsx",
        "November":  "LAB REPORT November 2023.xlsx",
        "December":  "LAB REPORT December 2023.xlsx",
    },
    2024: {
        "January":   "LAB REPORT January 2024.xlsx",
        "February":  "LAB REPORT February 2024.xlsx",
        "March":     "LAB REPORT March 2024.xlsx",
        "April":     "Lab report for April 2024.xlsx",
        "May":       "LAB REPORT May 2024.xlsx",
        "June":      "LAB REPORT June 2024.xlsx",
        "July":      "LAB REPORT July 2024.xlsx",
        "August":    "LAB REPORT August 2024.xlsx",
        "September": "LAB REPORT September 2024.xlsx",
        "October":   "LAB REPORT October 2024.xlsx",
        "November":  "LAB REPORT November 2024.xlsx",
        "December":  "LAB REPORT December 2024.xlsx",
    },
    2025: {
        "January":   "LAB REPORT January 2025.xlsx",
        "February":  "LAB REPORT February 2025.xlsx",
        "March":     "LAB REPORT March 2025.xlsx",
        "April":     "LAB REPORT April 2025.xlsx",
        "May":       "LAB REPORT MAY 2025.xlsx",
        "June":      "LAB REPORT JUNE 2025.xlsx",
        "July":      "LAB REPORT JULY 2025.xlsx",
        "August":    "LAB REPORT AUGUST 2025.xlsx",
        "September": "LAB REPORT SEPT 2025.xlsx",
        "October":   "LAB REPORT OCT 2025.xlsx",
        "November":  "LAB REPORT NOV 2025.xlsx",
        "December":  "LAB REPORT DEC 2025.xlsx",
    },
}

# ── Full column schema (display_name, dict_key) ────────────────────────────────

COLUMNS = [
    ("Date",                                        "date"),
    # Power
    ("Power GE (KW)",                               "power_ge"),
    ("Power NEA (KW)",                              "power_nea"),
    ("Power Total (KW)",                            "power_total"),
    ("Power / Flow (KW/ML)",                        "power_per_flow"),
    # Operational
    ("Flow (MLD)",                                  "flow"),
    # Inlet - grab (core 4)
    ("Inlet pH (Grab)",                                 "inlet_ph"),
    ("Inlet BOD (mg/L, Grab)",                          "inlet_bod"),
    ("Inlet COD (mg/L, Grab)",                          "inlet_cod"),
    ("Inlet TSS (mg/L, Grab)",                          "inlet_tss"),
    # Inlet - extended grab (grab-sourced; sparse in 2022+)
    ("Inlet TKN (mg/L, Grab)",                          "inlet_tkn_grab"),
    ("Inlet O&G (mg/L, Grab)",                          "inlet_og_grab"),
    ("Inlet PO4/TP (mg/L, Grab)",                       "inlet_po4_grab"),
    ("Inlet Total Coliform (CFU/100ml, Grab)",          "inlet_total_coliform_grab"),
    ("Inlet Fecal Coliform (CFU/100ml, Grab)",          "inlet_fecal_coliform_grab"),
    # Inlet - extended composite (composite-sourced; dense in 2022+)
    ("Inlet NH3-N (mg/L, Composite)",                   "inlet_nh3_comp"),
    ("Inlet O&G (mg/L, Composite)",                     "inlet_og_comp"),
    ("Inlet PO4/TP (mg/L, Composite)",                  "inlet_po4_comp"),
    ("Inlet Total Coliform (CFU/100ml, Composite)",     "inlet_total_coliform_comp"),
    ("Inlet Fecal Coliform (CFU/100ml, Composite)",     "inlet_fecal_coliform_comp"),
    # Inlet - composite core (null for 2021)
    ("Inlet pH (Composite)",                            "inlet_ph_comp"),
    ("Inlet BOD (mg/L, Composite)",                     "inlet_bod_comp"),
    ("Inlet COD (mg/L, Composite)",                     "inlet_cod_comp"),
    ("Inlet TSS (mg/L, Composite)",                     "inlet_tss_comp"),
    # Grit classifier
    ("Grit Classifier TSS (mg/L)",                  "grit_tss"),
    # Primary clarifier
    ("Primary Clarifier pH",                        "primary_ph"),
    ("Primary TSS (mg/L)",                          "primary_tss"),
    ("Primary BOD (mg/L)",                          "primary_bod"),
    ("Primary COD (mg/L)",                          "primary_cod"),
    ("Primary Sludge Totalizer (m3)",               "primary_sludge_totalizer"),
    # Secondary clarifier
    ("Sec Clarifier pH",                            "secondary_ph"),
    ("Sec Clarifier TSS (mg/L)",                    "secondary_tss"),
    ("Sec Clarifier BOD (mg/L)",                    "secondary_bod"),
    ("Sec Clarifier COD (mg/L)",                    "secondary_cod"),
    ("Sec Clarifier RAS",                           "secondary_ras"),
    # Secondary sedimentation
    ("Sec Sed pH",                                  "sec_sed_ph"),
    ("Sec Sed TSS (mg/L)",                          "sec_sed_tss"),
    ("Sec Sed BOD (mg/L)",                          "sec_sed_bod"),
    ("Sec Sed COD (mg/L)",                          "sec_sed_cod"),
    ("Sec Sed RAS (New)",                           "sec_sed_ras_new"),
    # Aeration - existing tank
    ("Aeration pH (Existing)",                      "aer_exist_ph"),
    ("Aeration DO (mg/L, Existing)",                "aer_exist_do"),
    ("Aeration MLSS (mg/L, Existing)",              "aer_exist_mlss"),
    ("Aeration MLVSS (mg/L, Existing)",             "aer_exist_mlvss"),
    ("Aeration SV30 (ml/L, Existing)",              "aer_exist_sv30"),
    ("Aeration SVI (Existing)",                     "aer_exist_svi"),
    # Aeration - new tank
    ("Aeration pH (New)",                           "aer_new_ph"),
    ("Aeration DO (mg/L, New)",                     "aer_new_do"),
    ("Aeration MLSS (mg/L, New)",                   "aer_new_mlss"),
    ("Aeration MLVSS (mg/L, New)",                  "aer_new_mlvss"),
    ("Aeration SV30 (ml/L, New)",                   "aer_new_sv30"),
    ("Aeration SVI (New)",                          "aer_new_svi"),
    # Effluent - grab
    ("Effluent pH (Grab)",                          "effluent_ph"),
    ("Effluent BOD (mg/L, Grab)",                   "effluent_bod"),
    ("Effluent COD (mg/L, Grab)",                   "effluent_cod"),
    ("Effluent TSS (mg/L, Grab)",                   "effluent_tss"),
    ("Effluent FRC (mg/L)",                         "effluent_frc"),
    ("Effluent O&G (mg/L)",                         "effluent_og"),
    ("Effluent NH3-N (mg/L)",                       "effluent_nh3"),
    ("Effluent Total Coliform (CFU/100ml)",         "effluent_total_coliform"),
    ("Effluent Fecal Coliform (CFU/100ml)",         "effluent_fecal_coliform"),
    # Effluent - composite (null for 2021)
    ("Effluent pH (Composite)",                     "effluent_ph_comp"),
    ("Effluent BOD (mg/L, Composite)",              "effluent_bod_comp"),
    ("Effluent COD (mg/L, Composite)",              "effluent_cod_comp"),
    ("Effluent TSS (mg/L, Composite)",              "effluent_tss_comp"),
]

ALL_KEYS = [k for _, k in COLUMNS]


# ── Value parsing ──────────────────────────────────────────────────────────────

def parse_value(raw):
    """Convert a raw cell value to float/None. Mirrors process_years.py logic."""
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return None   # date cells in data columns → ignore
    if isinstance(raw, (int, float)):
        v = float(raw)
        return None if (v != v) else v   # NaN guard
    s = str(raw).strip()
    if s in ("_", "-", "", "BDL", "*BDL", "N/A", "n/a",
             "#DIV/0!", "#REF!", "#N/A", "#VALUE!", "#NAME?", "Nil"):
        return None
    # Scientific notation like 1.4×10^9
    m = re.match(r"^([0-9.]+)\s*[×xX]\s*10\^?(\d+)$", s)
    if m:
        return float(m.group(1)) * (10 ** int(m.group(2)))
    m = re.match(r"^([0-9.]+)[eE]\+?(\d+)$", s)
    if m:
        return float(m.group(1)) * (10 ** int(m.group(2)))
    try:
        return float(s)
    except ValueError:
        return None


# ── Column auto-detection ──────────────────────────────────────────────────────

def detect_columns(ws, month_label=""):
    """
    Auto-detect column positions from sheet headers (rows 4-6).

    Works for all years 2021-2025 which share the same row layout:
      Row 4: major section headers
      Row 5: sub-section headers
      Row 6: parameter names
      Row 8: unit row (used to detect power unit)
      Row 9+: daily data

    Returns (data_cols dict, power_unit str).
    data_cols maps dict_key → 1-based column number (None if not found).
    """
    r4 = [c for c in ws.iter_rows(min_row=4, max_row=4, values_only=True)][0]
    r5 = [c for c in ws.iter_rows(min_row=5, max_row=5, values_only=True)][0]
    r6 = [c for c in ws.iter_rows(min_row=6, max_row=6, values_only=True)][0]
    r8 = [c for c in ws.iter_rows(min_row=8, max_row=8, values_only=True)][0]

    # Power unit from row 8 col 2 (col B)
    power_unit = "KWH" if (r8[1] and "KWH" in str(r8[1]).upper()) else "MW"

    def find_section_col(row, must_contain, must_not_contain=()):
        """Return 1-based col of first cell satisfying must/must-not criteria."""
        for i, v in enumerate(row):
            if v is None:
                continue
            s = str(v).upper()
            if all(kw in s for kw in must_contain) and \
               not any(kw in s for kw in must_not_contain):
                return i + 1
        return None

    def scan_params(start_col, params, row, search_range=15):
        """
        Starting at start_col, scan `row` for the named parameters.
        params: list of (key, substrings_to_match)
        Returns dict of key → 1-based col.
        """
        result = {k: None for k, _ in params}
        remaining = list(params)
        for i in range(start_col - 1, min(start_col - 1 + search_range, len(row))):
            v = row[i]
            if v is None:
                continue
            s = str(v).upper().strip()
            for k, matchers in list(remaining):
                if all(m in s for m in matchers):
                    result[k] = i + 1
                    remaining.remove((k, matchers))
                    break
            if not remaining:
                break
        return result

    # ── Power (fixed columns 2-5, 6 for flow) ─────────────────────────────────
    # These are always in the same position for all years.
    power_cols = {
        "power_ge":       2,
        "power_nea":      3,
        "power_total":    4,
        "power_per_flow": 5,
        "flow":           6,
    }

    # ── Grab inlet ─────────────────────────────────────────────────────────────
    # r5: 'Raw Sewage' (no COMPOSITE, no FLOW)
    grab_inlet_col = find_section_col(r5, ("RAW SEWAGE",), ("COMPOSITE", "FLOW"))
    inlet_ph = inlet_bod = inlet_cod = inlet_tss = None
    if grab_inlet_col:
        p = scan_params(grab_inlet_col, [
            ("inlet_ph",  ("PH",)),
            ("inlet_bod", ("BOD",)),
            ("inlet_cod", ("COD",)),
            ("inlet_tss", ("TSS",)),
        ], r6)
        inlet_ph, inlet_bod, inlet_cod, inlet_tss = (
            p["inlet_ph"], p["inlet_bod"], p["inlet_cod"], p["inlet_tss"])

    # ── Composite inlet ────────────────────────────────────────────────────────
    comp_inlet_col = find_section_col(r5, ("RAW SEWAGE", "COMPOSITE"), ("FLOW",))
    inlet_ph_comp = inlet_bod_comp = inlet_cod_comp = inlet_tss_comp = None
    if comp_inlet_col:
        p = scan_params(comp_inlet_col, [
            ("inlet_ph_comp",  ("PH",)),
            ("inlet_bod_comp", ("BOD",)),
            ("inlet_cod_comp", ("COD",)),
            ("inlet_tss_comp", ("TSS",)),
        ], r6)
        inlet_ph_comp  = p["inlet_ph_comp"]
        inlet_bod_comp = p["inlet_bod_comp"]
        inlet_cod_comp = p["inlet_cod_comp"]
        inlet_tss_comp = p["inlet_tss_comp"]

    # ── Extended inlet fields: provenance-aware dual scan ─────────────────────
    # Each extended feature (TKN/NH3-N, O&G, PO4/TP, Coliforms) is stored in
    # TWO output columns: grab-sourced and composite-sourced.  The grab sub-block
    # is scanned separately from the composite sub-block so values from
    # composite measurements are never mixed into grab prediction datasets.
    #
    # Measurement type history (from lab report analysis):
    #   TKN/NH3-N, O&G : grab in 2021,  composite from 2022+
    #   PO4/TP, Coliform: grab in 2021-2022, composite from 2023+
    #
    # Boundary: use Grit Classifier column if present; fall back to Primary
    # Clarifier so the scan is never skipped due to a missing Grit section.
    grit_col = find_section_col(r4, ("GRIT",))
    prim_col_early = find_section_col(r4, ("PRIMARY", "CLARIFIER"))
    inlet_scan_end = grit_col or prim_col_early

    inlet_tkn_grab = inlet_og_grab = inlet_po4_grab = None
    inlet_total_coliform_grab = inlet_fecal_coliform_grab = None
    inlet_nh3_comp = inlet_og_comp = inlet_po4_comp = None
    inlet_total_coliform_comp = inlet_fecal_coliform_comp = None

    if inlet_scan_end:
        already_basic = {x for x in [inlet_ph, inlet_bod, inlet_cod, inlet_tss,
                                      inlet_ph_comp, inlet_bod_comp,
                                      inlet_cod_comp, inlet_tss_comp] if x}

        def _scan_extended(start_col, end_col, out):
            for i in range(start_col - 1, min(end_col - 1, len(r6))):
                if (i + 1) in already_basic:
                    continue
                v = r6[i]
                if v is None:
                    continue
                s = str(v).upper().strip()
                if (("TKN" in s or "NH3" in s or "AMMONI" in s)
                        and "TOTAL" not in s and "tkn" not in out):
                    out["tkn"] = i + 1
                elif ("O & G" in s or ("OIL" in s and "GREAS" in s)) and "og" not in out:
                    out["og"] = i + 1
                elif (s in ("TP", "PO4") or "PHOSPH" in s) and "po4" not in out:
                    out["po4"] = i + 1
                elif "TOTAL COLIFORM" in s and "total_col" not in out:
                    out["total_col"] = i + 1
                elif "FECAL" in s and "fecal_col" not in out:
                    out["fecal_col"] = i + 1

        # Determine the column extent of each sub-block
        if grab_inlet_col and comp_inlet_col:
            if grab_inlet_col < comp_inlet_col:
                g_start, g_end = grab_inlet_col, comp_inlet_col
                c_start, c_end = comp_inlet_col, inlet_scan_end
            else:
                c_start, c_end = comp_inlet_col, grab_inlet_col
                g_start, g_end = grab_inlet_col, inlet_scan_end
        elif grab_inlet_col:
            g_start, g_end = grab_inlet_col, inlet_scan_end
            c_start, c_end = None, None
        else:
            g_start, g_end = None, None
            c_start, c_end = (comp_inlet_col, inlet_scan_end) if comp_inlet_col else (None, None)

        if g_start:
            gd = {}
            _scan_extended(g_start, g_end, gd)
            inlet_tkn_grab            = gd.get("tkn")
            inlet_og_grab             = gd.get("og")
            inlet_po4_grab            = gd.get("po4")
            inlet_total_coliform_grab = gd.get("total_col")
            inlet_fecal_coliform_grab = gd.get("fecal_col")

        if c_start:
            cd = {}
            _scan_extended(c_start, c_end, cd)
            inlet_nh3_comp            = cd.get("tkn")
            inlet_og_comp             = cd.get("og")
            inlet_po4_comp            = cd.get("po4")
            inlet_total_coliform_comp = cd.get("total_col")
            inlet_fecal_coliform_comp = cd.get("fecal_col")

    # ── Grit Classifier TSS ────────────────────────────────────────────────────
    # The section header column in r4 is at the same column as the TSS data.
    grit_tss = grit_col  # None if Grit section absent - handled downstream as missing value

    # ── Primary Clarifier ──────────────────────────────────────────────────────
    # prim_col_early was already detected above for the inlet scan fallback boundary.
    prim_col = prim_col_early
    primary_ph = primary_tss = primary_bod = primary_cod = primary_sludge_totalizer = None
    if prim_col:
        p = scan_params(prim_col, [
            ("primary_ph",  ("PH",)),
            ("primary_tss", ("TSS",)),
            ("primary_bod", ("BOD",)),
            ("primary_cod", ("COD",)),
        ], r6, search_range=8)
        primary_ph  = p["primary_ph"]
        primary_tss = p["primary_tss"]
        primary_bod = p["primary_bod"]
        primary_cod = p["primary_cod"]
        # Sludge Totalizer is just after COD
        if primary_cod:
            for i in range(primary_cod, min(primary_cod + 3, len(r6))):
                v = r6[i]
                if v and ("SLUDGE" in str(v).upper() or "TOTALIZER" in str(v).upper()):
                    primary_sludge_totalizer = i + 1
                    break

    # ── Secondary Clarifier ────────────────────────────────────────────────────
    sec_col = find_section_col(r4, ("SECONDARY", "CLARIFIER"), ("SEDIMENTATION",))
    secondary_ph = secondary_tss = secondary_bod = secondary_cod = secondary_ras = None
    if sec_col:
        p = scan_params(sec_col, [
            ("secondary_ph",  ("PH",)),
            ("secondary_tss", ("TSS",)),
            ("secondary_bod", ("BOD",)),
            ("secondary_cod", ("COD",)),
        ], r6, search_range=6)
        secondary_ph  = p["secondary_ph"]
        secondary_tss = p["secondary_tss"]
        secondary_bod = p["secondary_bod"]
        secondary_cod = p["secondary_cod"]
        if secondary_cod:
            # Scan up to 3 cols after COD for an explicit RAS header
            secondary_ras = None
            for i in range(secondary_cod, min(secondary_cod + 3, len(r6))):
                v = r6[i]
                if v and "RAS" in str(v).upper():
                    secondary_ras = i + 1
                    break
            if secondary_ras is None:
                secondary_ras = secondary_cod + 1  # fallback: column immediately after COD

    # ── Secondary Sedimentation ────────────────────────────────────────────────
    secsed_col = find_section_col(r4, ("SECONDARY", "SEDIMENTATION"))
    sec_sed_ph = sec_sed_tss = sec_sed_bod = sec_sed_cod = sec_sed_ras_new = None
    if secsed_col:
        p = scan_params(secsed_col, [
            ("sec_sed_ph",  ("PH",)),
            ("sec_sed_tss", ("TSS",)),
            ("sec_sed_bod", ("BOD",)),
            ("sec_sed_cod", ("COD",)),
        ], r6, search_range=6)
        sec_sed_ph  = p["sec_sed_ph"]
        sec_sed_tss = p["sec_sed_tss"]
        sec_sed_bod = p["sec_sed_bod"]
        sec_sed_cod = p["sec_sed_cod"]
        if sec_sed_cod:
            # Scan up to 3 cols after COD for an explicit RAS header
            sec_sed_ras_new = None
            for i in range(sec_sed_cod, min(sec_sed_cod + 3, len(r6))):
                v = r6[i]
                if v and "RAS" in str(v).upper():
                    sec_sed_ras_new = i + 1
                    break
            if sec_sed_ras_new is None:
                sec_sed_ras_new = sec_sed_cod + 1  # fallback: column immediately after COD

    # ── Aeration Tanks ─────────────────────────────────────────────────────────
    # Layout (same for all years):
    #   r4: 'Aeration Tank'  section header
    #   r5: 'Existing Aeration Tank' | 'New Aeration Tank'  subsection headers
    #   r6: PH  DO  MLSS  MLVSS  SV30  SVI  (×2 for each sub-tank)
    aer_main_col = find_section_col(r4, ("AERATION",))
    aer_exist_ph = aer_exist_do = aer_exist_mlss = aer_exist_mlvss = None
    aer_exist_sv30 = aer_exist_svi = None
    aer_new_ph = aer_new_do = aer_new_mlss = aer_new_mlvss = None
    aer_new_sv30 = aer_new_svi = None

    if aer_main_col:
        # Find subsection start columns in r5
        aer_exist_start = aer_new_start = None
        for i in range(aer_main_col - 1, min(aer_main_col + 25, len(r5))):
            v = r5[i]
            if v is None:
                continue
            s = str(v).upper()
            if "EXISTING" in s and "AERAT" in s and aer_exist_start is None:
                aer_exist_start = i + 1
            elif "NEW" in s and "AERAT" in s and aer_new_start is None:
                aer_new_start = i + 1

        def extract_aer_params(start_col):
            """Scan r6 from start_col for PH, DO, MLSS, MLVSS, SV30, SVI."""
            ph = do_ = mlss = mlvss = sv30 = svi = None
            if start_col is None:
                return ph, do_, mlss, mlvss, sv30, svi
            for i in range(start_col - 1, min(start_col + 10, len(r6))):
                v = r6[i]
                if v is None:
                    continue
                s = str(v).upper().strip()
                if s == "PH" and ph is None:
                    ph = i + 1
                elif s == "DO" and do_ is None:
                    do_ = i + 1
                elif s == "MLSS" and mlss is None:
                    mlss = i + 1
                elif s == "MLVSS" and mlvss is None:
                    mlvss = i + 1
                elif s == "SV30" and sv30 is None:
                    sv30 = i + 1
                elif s == "SVI" and svi is None:
                    svi = i + 1
                if all([ph, do_, mlss, mlvss, sv30, svi]):
                    break
            return ph, do_, mlss, mlvss, sv30, svi

        (aer_exist_ph, aer_exist_do, aer_exist_mlss,
         aer_exist_mlvss, aer_exist_sv30, aer_exist_svi) = extract_aer_params(aer_exist_start)

        (aer_new_ph, aer_new_do, aer_new_mlss,
         aer_new_mlvss, aer_new_sv30, aer_new_svi) = extract_aer_params(aer_new_start)

    # ── Effluent - grab CCT ────────────────────────────────────────────────────
    # r4: 'CHLORINE CONTACT TANK' without 'COMPOSITE', 'OUTLET', 'FLOW'
    grab_cct_col = find_section_col(r4, ("CHLORINE",), ("COMPOSITE", "OUTLET", "FLOW"))
    eff_ph = eff_bod = eff_cod = eff_tss = eff_frc = None
    if grab_cct_col:
        p = scan_params(grab_cct_col, [
            ("eff_ph",  ("PH",)),
            ("eff_bod", ("BOD",)),
            ("eff_cod", ("COD",)),
            ("eff_tss", ("TSS",)),
            ("eff_frc", ("FRC",)),
        ], r6, search_range=10)
        eff_ph  = p["eff_ph"]
        eff_bod = p["eff_bod"]
        eff_cod = p["eff_cod"]
        eff_tss = p["eff_tss"]
        eff_frc = p["eff_frc"]

    # ── Effluent - composite CCT ───────────────────────────────────────────────
    comp_cct_col = find_section_col(r4, ("CHLORINE", "COMPOSITE"))
    eff_ph_comp = eff_bod_comp = eff_cod_comp = eff_tss_comp = None
    if comp_cct_col:
        p = scan_params(comp_cct_col, [
            ("eff_ph_comp",  ("PH",)),
            ("eff_bod_comp", ("BOD",)),
            ("eff_cod_comp", ("COD",)),
            ("eff_tss_comp", ("TSS",)),
        ], r6, search_range=10)
        eff_ph_comp  = p["eff_ph_comp"]
        eff_bod_comp = p["eff_bod_comp"]
        eff_cod_comp = p["eff_cod_comp"]
        eff_tss_comp = p["eff_tss_comp"]

    # ── Extended effluent fields (O&G, NH3, coliforms) ────────────────────────
    # These are in the CCT OUTLET FLOW section (2022+) or within the CCT grab
    # section (2021). Scanning from the earliest CCT column captures both cases.
    effluent_og = effluent_nh3 = effluent_total_coliform = effluent_fecal_coliform = None
    cct_cols = [x for x in [grab_cct_col, comp_cct_col] if x]
    if cct_cols:
        # Also check for a CCT OUTLET FLOW section in r4 (2022+)
        outlet_flow_col = find_section_col(r4, ("CHLORINE", "OUTLET", "FLOW"),
                                           ("COMPOSITE",))
        all_cct = [x for x in [grab_cct_col, comp_cct_col, outlet_flow_col] if x]
        scan_start = min(all_cct)
        scan_end = max(all_cct) + 25
        for i in range(scan_start - 1, min(scan_end, len(r6))):
            v = r6[i]
            if v is None:
                continue
            s = str(v).upper().strip()
            if ("O & G" in s or ("OIL" in s and "GREAS" in s)) and effluent_og is None:
                effluent_og = i + 1
            elif ("AMMONI" in s or s == "NH3-N") and effluent_nh3 is None:
                effluent_nh3 = i + 1
            elif "TOTAL COLIFORM" in s and effluent_total_coliform is None:
                effluent_total_coliform = i + 1
            elif "FECAL" in s and effluent_fecal_coliform is None:
                effluent_fecal_coliform = i + 1

    # ── Build final column map ─────────────────────────────────────────────────
    data_cols = {
        **power_cols,
        "inlet_ph":                  inlet_ph,
        "inlet_bod":                 inlet_bod,
        "inlet_cod":                 inlet_cod,
        "inlet_tss":                 inlet_tss,
        "inlet_tkn_grab":             inlet_tkn_grab,
        "inlet_og_grab":             inlet_og_grab,
        "inlet_po4_grab":            inlet_po4_grab,
        "inlet_total_coliform_grab": inlet_total_coliform_grab,
        "inlet_fecal_coliform_grab": inlet_fecal_coliform_grab,
        "inlet_nh3_comp":            inlet_nh3_comp,
        "inlet_og_comp":             inlet_og_comp,
        "inlet_po4_comp":            inlet_po4_comp,
        "inlet_total_coliform_comp": inlet_total_coliform_comp,
        "inlet_fecal_coliform_comp": inlet_fecal_coliform_comp,
        "inlet_ph_comp":             inlet_ph_comp,
        "inlet_bod_comp":            inlet_bod_comp,
        "inlet_cod_comp":            inlet_cod_comp,
        "inlet_tss_comp":            inlet_tss_comp,
        "grit_tss":                  grit_tss,
        "primary_ph":                primary_ph,
        "primary_tss":               primary_tss,
        "primary_bod":               primary_bod,
        "primary_cod":               primary_cod,
        "primary_sludge_totalizer":  primary_sludge_totalizer,
        "secondary_ph":              secondary_ph,
        "secondary_tss":             secondary_tss,
        "secondary_bod":             secondary_bod,
        "secondary_cod":             secondary_cod,
        "secondary_ras":             secondary_ras,
        "sec_sed_ph":                sec_sed_ph,
        "sec_sed_tss":               sec_sed_tss,
        "sec_sed_bod":               sec_sed_bod,
        "sec_sed_cod":               sec_sed_cod,
        "sec_sed_ras_new":           sec_sed_ras_new,
        "aer_exist_ph":              aer_exist_ph,
        "aer_exist_do":              aer_exist_do,
        "aer_exist_mlss":            aer_exist_mlss,
        "aer_exist_mlvss":           aer_exist_mlvss,
        "aer_exist_sv30":            aer_exist_sv30,
        "aer_exist_svi":             aer_exist_svi,
        "aer_new_ph":                aer_new_ph,
        "aer_new_do":                aer_new_do,
        "aer_new_mlss":              aer_new_mlss,
        "aer_new_mlvss":             aer_new_mlvss,
        "aer_new_sv30":              aer_new_sv30,
        "aer_new_svi":               aer_new_svi,
        "effluent_ph":               eff_ph,
        "effluent_bod":              eff_bod,
        "effluent_cod":              eff_cod,
        "effluent_tss":              eff_tss,
        "effluent_frc":              eff_frc,
        "effluent_og":               effluent_og,
        "effluent_nh3":              effluent_nh3,
        "effluent_total_coliform":   effluent_total_coliform,
        "effluent_fecal_coliform":   effluent_fecal_coliform,
        "effluent_ph_comp":          eff_ph_comp,
        "effluent_bod_comp":         eff_bod_comp,
        "effluent_cod_comp":         eff_cod_comp,
        "effluent_tss_comp":         eff_tss_comp,
    }

    # Diagnostic: report any undetected critical columns
    critical = ["inlet_ph", "inlet_bod", "inlet_cod", "inlet_tss",
                "secondary_ph", "secondary_tss",
                "effluent_ph", "effluent_bod", "effluent_cod", "effluent_tss"]
    missing = [k for k in critical if data_cols.get(k) is None]
    if missing:
        print(f"    WARNING {month_label}: could not detect columns for: {missing}")

    aer_detected = [k for k in ["aer_exist_do", "aer_exist_mlss", "aer_new_do", "aer_new_mlss"]
                    if data_cols.get(k) is not None]
    if not aer_detected:
        print(f"    WARNING {month_label}: aeration tank columns not detected")

    return data_cols, power_unit


# ── Month extraction ───────────────────────────────────────────────────────────

def extract_month(year, month_name, filename):
    """
    Read one monthly Excel file and return list of daily dicts.
    Returns empty list if the file does not exist.
    """
    filepath = os.path.join(BASE_DIR, "raw_data", str(year), filename)
    if not os.path.exists(filepath):
        print(f"    WARNING: {filename} not found - skipping.")
        return []

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]

    label = f"{year}/{month_name}"
    data_cols, raw_power_unit = detect_columns(ws, label)

    days = []
    for row_idx in range(DATA_START_ROW, DATA_START_ROW + 65):
        date_cell = ws.cell(row=row_idx, column=1).value

        # Stop at end of data
        if date_cell is None:
            break
        if isinstance(date_cell, str):
            s = date_cell.strip().upper()
            if s in ("AVG", "REMARKS:", ""):
                break
            if s == "AVERAGE":
                continue   # weekly average row - skip, keep reading
            break           # any other string → end of data
        if not isinstance(date_cell, datetime):
            break

        day = {"date": date_cell.strftime("%Y-%m-%d")}
        for key, col in data_cols.items():
            day[key] = parse_value(ws.cell(row=row_idx, column=col).value) if col else None

        # Normalise power units: MW → KWh (×1000)
        if raw_power_unit == "MW":
            for pk in POWER_KEYS:
                if day.get(pk) is not None:
                    day[pk] = round(day[pk] * 1000, 3)

        days.append(day)

    return days


# ── CSV loader for 2020 data ───────────────────────────────────────────────────

def load_csv(csv_path):
    """
    Load 2020 RAW.csv.
    CSV columns: Date, Q_IN, IN_BOD, IN_COD, IN_TSS, IN_pH, PC, PF, O_BOD, O_COD, O_TSS
    PC is NEA power in MWh → ×1000 to get KWh.
    PF (power/flow) is already in KW/ML.
    All aeration/intermediate/quality fields are left as None.
    """
    rows = []
    with open(csv_path, newline="") as f:
        reader = csv.DictReader(f)
        for rec in reader:
            def v(col):
                raw = rec.get(col, "").strip()
                if raw in ("", "-", "N/A"):
                    return None
                try:
                    return float(raw)
                except ValueError:
                    return None

            date_str = rec.get("Date", "").strip()
            try:
                dt = datetime.strptime(date_str, "%m/%d/%Y")
                date_fmt = dt.strftime("%Y-%m-%d")
            except ValueError:
                continue   # skip malformed dates

            pc_mwh = v("PC")
            pc_kw  = pc_mwh * 1000 if pc_mwh is not None else None

            row = {k: None for k in ALL_KEYS}
            row.update({
                "date":           date_fmt,
                "power_ge":       None,      # not in CSV
                "power_nea":      pc_kw,
                "power_total":    pc_kw,
                "power_per_flow": v("PF"),
                "flow":           v("Q_IN"),
                "inlet_ph":       v("IN_pH"),
                "inlet_bod":      v("IN_BOD"),
                "inlet_cod":      v("IN_COD"),
                "inlet_tss":      v("IN_TSS"),
                "effluent_bod":   v("O_BOD"),
                "effluent_cod":   v("O_COD"),
                "effluent_tss":   v("O_TSS"),
            })
            rows.append(row)
    return rows


# ── Main merge ─────────────────────────────────────────────────────────────────

def collect_all_rows():
    """
    Collect all daily rows in chronological order.
    Priority: Excel-extracted rows > CSV rows for any overlapping dates.
    """
    excel_rows = []
    for year in sorted(YEAR_FILES):
        print(f"  {year}:")
        year_count = 0
        for month_name, filename in YEAR_FILES[year].items():
            days = extract_month(year, month_name, filename)
            excel_rows.extend(days)
            year_count += len(days)
            print(f"    {month_name}: {len(days)} days")
        print(f"    → {year_count} rows")

    excel_dates = {r["date"] for r in excel_rows}

    csv_path = os.path.join(BASE_DIR, "raw_data", "2020", "RAW.csv")
    csv_rows = load_csv(csv_path)
    csv_kept = [r for r in csv_rows if r["date"] not in excel_dates]
    csv_skip = len(csv_rows) - len(csv_kept)
    print(f"  CSV: {len(csv_rows)} rows total, "
          f"{csv_skip} skipped (overlap with Excel), {len(csv_kept)} kept")

    all_rows = csv_kept + excel_rows
    all_rows.sort(key=lambda r: r.get("date", ""))

    # Ensure every row has all keys (fill missing with None)
    for row in all_rows:
        for k in ALL_KEYS:
            row.setdefault(k, None)

    return all_rows


# ── Excel writing ──────────────────────────────────────────────────────────────

def _side(s="thin"):
    return Side(style=s)

def thin_border():
    s = _side()
    return Border(left=s, right=s, top=s, bottom=s)

# Section colour palette
_FILLS = {
    "date":       PatternFill("solid", fgColor="1F4E79"),
    "power":      PatternFill("solid", fgColor="2E75B6"),
    "flow":       PatternFill("solid", fgColor="2E75B6"),
    "inlet":      PatternFill("solid", fgColor="375623"),
    "grit":       PatternFill("solid", fgColor="7030A0"),
    "primary":    PatternFill("solid", fgColor="7030A0"),
    "secondary":  PatternFill("solid", fgColor="006B6B"),
    "sec_sed":    PatternFill("solid", fgColor="2E4057"),
    "aer":        PatternFill("solid", fgColor="7B3F00"),   # dark brown/amber
    "effluent":   PatternFill("solid", fgColor="833C00"),
}

WHITE_BOLD  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
NORMAL_FONT = Font(name="Calibri", size=10)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center")

def col_fill(jkey):
    if jkey == "date":               return _FILLS["date"]
    if jkey.startswith("power") or jkey == "flow":  return _FILLS["power"]
    if jkey.startswith("inlet"):     return _FILLS["inlet"]
    if jkey.startswith("grit"):      return _FILLS["grit"]
    if jkey.startswith("primary"):   return _FILLS["primary"]
    if jkey.startswith("secondary"): return _FILLS["secondary"]
    if jkey.startswith("sec_sed"):   return _FILLS["sec_sed"]
    if jkey.startswith("aer"):       return _FILLS["aer"]
    if jkey.startswith("effluent"):  return _FILLS["effluent"]
    return _FILLS["date"]


def write_excel(rows, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "All Years 2020-2025 Full"
    ws.freeze_panes = "B2"

    num_cols = len(COLUMNS)

    # Header row
    for ci, (label, jkey) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=ci, value=label)
        cell.fill   = col_fill(jkey)
        cell.font   = WHITE_BOLD
        cell.alignment = CENTER
        cell.border = thin_border()
    ws.row_dimensions[1].height = 40

    # Data rows
    for ri, day in enumerate(rows, start=2):
        for ci, (_, jkey) in enumerate(COLUMNS, start=1):
            raw = day.get(jkey)
            if jkey == "date" and raw:
                parts = raw.split("-")
                try:
                    value = dtdate(int(parts[0]), int(parts[1]), int(parts[2]))
                    fmt   = "DD-MMM-YYYY"
                except Exception:
                    value, fmt = raw, None
            elif raw is None:
                value, fmt = "", None
            elif isinstance(raw, float):
                value, fmt = round(raw, 4), None
            else:
                value, fmt = raw, None

            cell = ws.cell(row=ri, column=ci, value=value)
            cell.font      = NORMAL_FONT
            cell.alignment = LEFT if jkey == "date" else CENTER
            cell.border    = thin_border()
            if jkey == "date" and fmt:
                cell.number_format = fmt

    # Column widths (approximate, by section)
    widths = [
        13,                          # Date
        11, 11, 12, 14, 9,           # Power x4, Flow
        9, 13, 13, 13,               # Inlet grab (pH, BOD, COD, TSS)
        14, 14, 14, 18, 18,          # Inlet extended grab (TKN, O&G, PO4, Coliform x2)
        16, 16, 16, 22, 22,          # Inlet extended composite (NH3-N, O&G, PO4, Coliform x2)
        13, 17, 17, 17,              # Inlet composite core
        14,                          # Grit TSS
        13, 13, 13, 13, 15,          # Primary (pH, TSS, BOD, COD, Sludge)
        13, 15, 15, 15, 13,          # Sec Clarifier
        11, 14, 14, 14, 12,          # Sec Sed
        14, 16, 16, 16, 14, 12,      # Aeration existing (pH, DO, MLSS, MLVSS, SV30, SVI)
        14, 14, 14, 14, 14, 12,      # Aeration new
        12, 16, 16, 16, 11,          # Effluent grab (pH, BOD, COD, TSS, FRC)
        13, 15, 18, 18,              # Effluent O&G, NH3, Coliform x2
        16, 20, 20, 20,              # Effluent composite
    ]
    for i, w in enumerate(widths[:num_cols], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(output_path)


# ── Entry point ────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("Extracting full dataset (2020-2025)...\n")
    rows = collect_all_rows()
    output = os.path.join(BASE_DIR, "raw_data", "All_Years_Full.xlsx")
    print(f"\nWriting {len(rows)} rows to {output}...")
    write_excel(rows, output)
    print(f"Done. Total rows: {len(rows)}, columns: {len(COLUMNS)}")

    # Summary statistics
    import collections
    year_counts = collections.Counter(r["date"][:4] for r in rows)
    print("\nRows per year:")
    for yr in sorted(year_counts):
        print(f"  {yr}: {year_counts[yr]}")

    # Non-null rate for key new columns
    new_cols = ["inlet_tkn", "inlet_og", "grit_tss", "primary_ph",
                "aer_exist_do", "aer_exist_mlss", "aer_exist_svi",
                "aer_new_do", "aer_new_mlss", "effluent_og", "effluent_nh3"]
    print("\nCoverage for new columns:")
    n = len(rows)
    for k in new_cols:
        filled = sum(1 for r in rows if r.get(k) is not None)
        print(f"  {k:<30} {filled:4d}/{n}  ({filled/n*100:.0f}%)")
